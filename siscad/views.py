from datetime import date, datetime, timedelta
from django.utils import timezone
from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse
from django.contrib import messages
from django.db import transaction
from django.db.models import Q, Count
from .forms import UploadExcelForm
from django.db.models import Avg, Max, Min, Count, Q
import io


import openpyxl
from openpyxl.utils import get_column_letter
from .models import (
    Profesor,
    Alumno,
    Secretaria,
    Administrador,
    Nota,
    Curso,
    MatriculaCurso,
    GrupoTeoria,
    GrupoLaboratorio,
    MatriculaLaboratorio,
    Hora,
    AsistenciaAlumno,
    GrupoPractica,
    AsistenciaProfesor,
    Reserva,
    Aula,
    Examen,
)
import pandas as pd


def inicio(request):
    nombre = request.session.get("nombre")
    rol = request.session.get("rol")
    if not nombre:
        return redirect("login")

    return render(request, "siscad/inicio.html", {"nombre": nombre, "rol": rol})


def login_view(request):
    if request.method == "POST":
        email = request.POST.get("email")
        dni = request.POST.get("dni")

        usuario = None
        rol = None

        if Profesor.objects.filter(email=email, dni=dni).exists():
            usuario = Profesor.objects.get(email=email, dni=dni)
            rol = "Profesor"
        elif Alumno.objects.filter(email=email, dni=dni).exists():
            usuario = Alumno.objects.get(email=email, dni=dni)
            rol = "Alumno"
        elif Secretaria.objects.filter(email=email, dni=dni).exists():
            usuario = Secretaria.objects.get(email=email, dni=dni)
            rol = "Secretaria"
        elif Administrador.objects.filter(email=email, dni=dni).exists():
            usuario = Administrador.objects.get(email=email, dni=dni)
            rol = "Administrador"

        if usuario:
            request.session["usuario_id"] = usuario.id
            request.session["rol"] = rol
            request.session["nombre"] = usuario.nombre
            request.session["email"] = email
            messages.success(request, f"Bienvenido {rol} {usuario.nombre}")

            return redirect(f"inicio_{rol.lower()}")

        else:
            messages.error(request, "Email o DNI incorrectos")

    return render(request, "siscad/login.html")


def logout_view(request):
    request.session.flush()
    return redirect("login")


# =======================Vista de Secretaria ===============================================
def inicio_secretaria(request):
    rol = request.session.get("rol")

    if rol != "Secretaria":
        request.session["rol"] = "Ninguno"
        return redirect("login")

    nombre = request.session.get("nombre")
    return render(
        request, "siscad/secretaria/menu.html", {"nombre": nombre, "rol": rol}
    )


def insertar_alumnos_excel(request):
    if request.method == "POST":
        form = UploadExcelForm(request.POST, request.FILES)
        if form.is_valid():
            file = form.cleaned_data["file"]

            try:
                df = pd.read_excel(file)
                df.columns = [str(col).strip().lower() for col in df.columns]

                required_columns = [
                    "apellidop",
                    "apellidom",
                    "nombres",
                    "correo",
                    "dni",
                    "cui",
                ]
                missing = [col for col in required_columns if col not in df.columns]
                if missing:
                    messages.error(
                        request, f"Faltan columnas obligatorias: {', '.join(missing)}"
                    )
                    return redirect("insertar_alumnos_excel")

                created = 0
                updated = 0
                errores = []

                with transaction.atomic():
                    for index, row in df.iterrows():
                        apellidop = (
                            str(row["apellidop"]).strip()
                            if pd.notna(row["apellidop"])
                            else ""
                        )
                        apellidom = (
                            str(row["apellidom"]).strip()
                            if pd.notna(row["apellidom"])
                            else ""
                        )
                        nombres = (
                            str(row["nombres"]).strip()
                            if pd.notna(row["nombres"])
                            else ""
                        )
                        email = (
                            str(row["correo"]).strip().lower()
                            if pd.notna(row["correo"])
                            else ""
                        )
                        dni = str(row["dni"]).strip() if pd.notna(row["dni"]) else ""
                        cui = str(row["cui"]).strip() if pd.notna(row["cui"]) else ""

                        nombre = f"{apellidop} {apellidom} {nombres}".strip()

                        if not email:
                            errores.append(
                                f"Fila {index + 2}: email vacío, no se pudo procesar."
                            )
                            continue

                        alumno, created_flag = Alumno.objects.update_or_create(
                            email=email,
                            defaults={"nombre": nombre, "dni": dni, "cui": cui},
                        )

                        created += 1 if created_flag else updated + 1

                messages.success(
                    request,
                    f" Importación completada: {created} creados, {updated} actualizados.",
                )
                if errores:
                    for err in errores[:10]:
                        messages.warning(request, err)

                return redirect("insertar_alumnos_excel")

            except Exception as e:
                messages.error(request, f" Error leyendo el archivo: {e}")
                return redirect("insertar_alumnos_excel")

    else:
        form = UploadExcelForm()

    alumnos = Alumno.objects.all()

    return render(
        request,
        "siscad/secretaria/insertar_alumnos_excel.html",
        {"form": form, "alumnos": alumnos},
    )


def listar_alumno_grupo_teoria(request):
    alumnos = []
    curso_id = request.POST.get("curso_id", "")
    turno = request.POST.get("turno", "")
    semestre_tipo = request.POST.get("semestre_tipo", "")

    # Filtrar cursos según semestre seleccionado
    todos_cursos = Curso.objects.all()
    if semestre_tipo == "par":
        cursos = [c for c in todos_cursos if c.semestre % 2 == 0]
    elif semestre_tipo == "impar":
        cursos = [c for c in todos_cursos if c.semestre % 2 != 0]
    else:
        cursos = todos_cursos

    # Obtener turnos disponibles y alumnos si se seleccionó curso
    turnos = []
    if request.method == "POST" and curso_id:
        turnos = (
            GrupoTeoria.objects.filter(curso_id=curso_id)
            .values_list("turno", flat=True)
            .distinct()
        )

        if turno:
            matriculas = MatriculaCurso.objects.filter(curso_id=curso_id, turno=turno)
            alumnos = [matricula.alumno for matricula in matriculas]

        # Descargar Excel
        if "descargar_excel" in request.POST and alumnos:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Alumnos"

            # Encabezados
            headers = ["Nombre", "Email", "DNI", "CUI"]
            for col_num, header in enumerate(headers, 1):
                ws[f"{get_column_letter(col_num)}1"] = header

            # Datos de alumnos
            for row_num, alumno in enumerate(alumnos, 2):
                ws[f"A{row_num}"] = alumno.nombre
                ws[f"B{row_num}"] = alumno.email
                ws[f"C{row_num}"] = alumno.dni
                ws[f"D{row_num}"] = alumno.cui

            # Nombre del archivo con curso y turno
            curso_nombre = Curso.objects.get(id=curso_id).nombre
            archivo_nombre = f"Alumnos_{curso_nombre}_Turno_{turno}.xlsx"

            response = HttpResponse(
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            response["Content-Disposition"] = f'attachment; filename="{archivo_nombre}"'
            wb.save(response)
            return response

    return render(
        request,
        "siscad/secretaria/listar_alumnos_grupo_teoria.html",
        {
            "alumnos": alumnos,
            "cursos": cursos,
            "turnos": turnos,
            "curso_id": curso_id,
            "turno": turno,
            "semestre_tipo": semestre_tipo,
        },
    )


def listar_grupos_laboratorio(request):
    laboratorios = GrupoLaboratorio.objects.select_related(
        "grupo_teoria", "grupo_teoria__curso", "profesor"
    ).all()

    if request.method == "POST":
        # Actualizar cupos de un laboratorio
        if "actualizar_cupos" in request.POST:
            lab_id = request.POST.get("lab_id")
            nuevos_cupos = request.POST.get("cupos")

            if lab_id and nuevos_cupos:
                try:
                    lab = GrupoLaboratorio.objects.get(id=lab_id)
                    lab.cupos = int(nuevos_cupos)
                    lab.save()
                    messages.success(
                        request,
                        f"Cupos actualizados para {lab.grupo_teoria.curso.nombre} - Lab {lab.grupo}",
                    )
                    return redirect("listar_grupos_laboratorio")
                except GrupoLaboratorio.DoesNotExist:
                    messages.error(request, "Laboratorio no encontrado")
                except ValueError:
                    messages.error(request, "El número de cupos debe ser un entero")

        # Descargar Excel
        elif "descargar_excel" in request.POST:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Laboratorios"

            headers = [
                "Curso",
                "Grupo Teoría",
                "Grupo Laboratorio",
                "Profesor",
                "Cupos",
            ]
            for col_num, header in enumerate(headers, 1):
                ws[f"{get_column_letter(col_num)}1"] = header

            for row_num, lab in enumerate(laboratorios, 2):
                ws[f"A{row_num}"] = lab.grupo_teoria.curso.nombre
                ws[f"B{row_num}"] = lab.grupo_teoria.turno
                ws[f"C{row_num}"] = lab.grupo
                ws[f"D{row_num}"] = (
                    lab.profesor.nombre if lab.profesor else "Sin asignar"
                )
                ws[f"E{row_num}"] = lab.cupos

            archivo_nombre = "Laboratorios_Disponibles.xlsx"
            response = HttpResponse(
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            response["Content-Disposition"] = f'attachment; filename="{archivo_nombre}"'
            wb.save(response)
            return response

    return render(
        request,
        "siscad/secretaria/listar_grupos_laboratorio.html",
        {"laboratorios": laboratorios},
    )


def listar_alumno_grupo_laboratorio(request):
    cursos = Curso.objects.all()
    laboratorios = []
    alumnos = []
    curso_id = None
    lab_id = None

    if request.method == "POST":
        curso_id = request.POST.get("curso_id")
        lab_id = request.POST.get("lab_id")

        # Filtrar laboratorios por curso seleccionado
        if curso_id:
            laboratorios = GrupoLaboratorio.objects.filter(
                grupo_teoria__curso_id=curso_id
            )

        # Buscar alumnos
        if lab_id and "buscar_alumnos" in request.POST:
            matriculas = MatriculaLaboratorio.objects.filter(
                grupo_laboratorio_id=lab_id
            )
            alumnos = [m.alumno for m in matriculas]

        # Descargar Excel
        if lab_id and "descargar_excel" in request.POST:
            matriculas = MatriculaLaboratorio.objects.filter(
                grupo_laboratorio_id=lab_id
            )
            alumnos = [m.alumno for m in matriculas]

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Alumnos Laboratorio"

            headers = ["Nombre", "Email", "DNI", "CUI"]
            for col_num, header in enumerate(headers, 1):
                ws[f"{get_column_letter(col_num)}1"] = header

            for row_num, alumno in enumerate(alumnos, 2):
                ws[f"A{row_num}"] = alumno.nombre
                ws[f"B{row_num}"] = alumno.email
                ws[f"C{row_num}"] = alumno.dni
                ws[f"D{row_num}"] = alumno.cui

            response = HttpResponse(
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            response["Content-Disposition"] = (
                'attachment; filename="Alumnos_Laboratorio.xlsx"'
            )
            wb.save(response)
            return response

    return render(
        request,
        "siscad/secretaria/listar_alumno_grupo_laboratorio.html",
        {
            "cursos": cursos,
            "laboratorios": laboratorios,
            "alumnos": alumnos,
            "curso_id": curso_id,
            "lab_id": lab_id,
        },
    )


def visualizar_horarios_aulas(request):
    dias = ["L", "M", "X", "J", "V"]
    dias_nombres = {
        "L": "Lunes",
        "M": "Martes",
        "X": "Miércoles",
        "J": "Jueves",
        "V": "Viernes",
    }

    dias_lista = [(clave, dias_nombres[clave]) for clave in dias]

    horarios = Hora.objects.select_related(
        "aula",
        "grupo_teoria__curso",
        "grupo_practica__grupo_teoria__curso",
        "grupo_laboratorio__grupo_teoria__curso",
    ).order_by("aula__nombre", "hora_inicio")

    tabla_horarios = {}

    for h in horarios:
        if not h.aula:
            continue

        aula = h.aula.nombre
        bloque = f"{h.hora_inicio.strftime('%H:%M')} - {h.hora_fin.strftime('%H:%M')}"
        dia = h.dia

        if aula not in tabla_horarios:
            tabla_horarios[aula] = {}

        if bloque not in tabla_horarios[aula]:
            tabla_horarios[aula][bloque] = {d: "" for d in dias}

        if h.grupo_teoria:
            curso = h.grupo_teoria.curso.nombre
            grupo = f"T{h.grupo_teoria.id}"
        elif h.grupo_practica:
            curso = h.grupo_practica.grupo_teoria.curso.nombre
            grupo = f"P{h.grupo_practica.id}"
        elif h.grupo_laboratorio:
            curso = h.grupo_laboratorio.grupo_teoria.curso.nombre
            grupo = f"L{h.grupo_laboratorio.id}"
        else:
            curso = "Receso"
            grupo = ""

        tabla_horarios[aula][bloque][dia] = f"{curso} ({grupo})"

    for aula in tabla_horarios:
        tabla_horarios[aula] = dict(
            sorted(tabla_horarios[aula].items(), key=lambda x: x[0])
        )

    context = {
        "dias_lista": dias_lista,
        "tabla_horarios": {
            aula: [
                {
                    "bloque": bloque,
                    "dias": [(dia, dias_data[dia]) for dia, _ in dias_lista],
                }
                for bloque, dias_data in bloques.items()
            ]
            for aula, bloques in tabla_horarios.items()
        },
    }

    return render(request, "siscad/secretaria/visualizar_horarios_aulas.html", context)


# =======================Vista de Alumno====================================================
def inicio_alumno(request):
    rol = request.session.get("rol")

    if rol != "Alumno":
        request.session["rol"] = "Ninguno"
        return redirect("login")

    nombre = request.session.get("nombre")
    return render(request, "siscad/alumno/menu.html", {"nombre": nombre, "rol": rol})


def visualizar_notas(request):
    rol = request.session.get("rol")

    if rol != "Alumno":
        request.session["rol"] = "Ninguno"
        return redirect("login")

    nombre = request.session.get("nombre")

    alumno = Alumno.objects.filter(nombre=nombre).first()

    if not alumno:
        return redirect("inicio_alumno")

    # Verificar si se solicita descargar Excel
    if request.method == "POST" and "descargar_excel" in request.POST:
        return descargar_libreta_excel(alumno)

    matriculas = MatriculaCurso.objects.filter(alumno=alumno).select_related("curso")

    notas_por_curso = []
    estadisticas_generales = {
        "total_cursos": 0,
        "cursos_aprobados": 0,
        "cursos_desaprobados": 0,
        "cursos_en_proceso": 0,
        "promedio_general": 0,
        "mejor_nota": 0,
        "peor_nota": 20,
        "total_notas_registradas": 0,
        "cursos_con_sustitutorio": 0,
        "total_notas_esperadas": 0,
        "progreso_general": 0,
    }

    suma_promedios = 0
    cursos_con_promedio = 0
    total_notas_registradas_global = 0
    total_notas_esperadas_global = 0

    for matricula in matriculas:
        curso = matricula.curso
        turno = matricula.turno

        # Obtener todas las notas del curso
        notas_curso = Nota.objects.filter(alumno=alumno, curso=curso).order_by(
            "tipo", "periodo"
        )

        # Organizar notas por tipo - VERSIÓN CON -1
        notas_parciales = {1: None, 2: None, 3: None}
        notas_continuas = {1: None, 2: None, 3: None}
        nota_sustitutorio = None

        for nota in notas_curso:
            if nota.tipo == "P" and nota.periodo in [1, 2, 3]:
                # Solo considerar valores >= 0
                notas_parciales[nota.periodo] = (
                    nota.valor if nota.valor is not None and nota.valor >= 0 else None
                )
            elif nota.tipo == "C" and nota.periodo in [1, 2, 3]:
                notas_continuas[nota.periodo] = (
                    nota.valor if nota.valor is not None and nota.valor >= 0 else None
                )
            elif nota.tipo == "S":
                nota_sustitutorio = (
                    nota.valor if nota.valor is not None and nota.valor >= 0 else None
                )

        # Calcular nota final considerando sustitutorio
        promedio_final = calcular_nota_final_alumno(alumno, curso)

        # Determinar estado del curso
        estado_curso = determinar_estado_curso(
            promedio_final, curso, notas_parciales, notas_continuas
        )

        # Estadísticas del curso - VERSIÓN CON -1
        notas_validas = [
            n.valor for n in notas_curso if n.valor is not None and n.valor >= 0
        ]
        total_notas_curso = len(notas_validas)
        promedio_curso = sum(notas_validas) / len(notas_validas) if notas_validas else 0

        # Calcular progreso del curso
        total_notas_esperadas = calcular_total_notas_esperadas(curso)
        progreso_curso = (
            (total_notas_curso / total_notas_esperadas * 100)
            if total_notas_esperadas > 0
            else 0
        )

        # Actualizar estadísticas generales
        estadisticas_generales["total_cursos"] += 1
        estadisticas_generales["total_notas_registradas"] += total_notas_curso
        total_notas_registradas_global += total_notas_curso
        total_notas_esperadas_global += total_notas_esperadas

        if estado_curso == "aprobado":
            estadisticas_generales["cursos_aprobados"] += 1
        elif estado_curso == "desaprobado":
            estadisticas_generales["cursos_desaprobados"] += 1
        else:
            estadisticas_generales["cursos_en_proceso"] += 1

        if promedio_final and promedio_final >= 0:
            suma_promedios += promedio_final
            cursos_con_promedio += 1
            estadisticas_generales["mejor_nota"] = max(
                estadisticas_generales["mejor_nota"], promedio_final
            )
            estadisticas_generales["peor_nota"] = min(
                estadisticas_generales["peor_nota"], promedio_final
            )

        if nota_sustitutorio is not None and nota_sustitutorio >= 0:
            estadisticas_generales["cursos_con_sustitutorio"] += 1

        # Preparar datos para mostrar en tabla
        datos_parciales = []
        datos_continuas = []

        for periodo in [1, 2, 3]:
            peso_parcial = getattr(curso, f"peso_parcial_{periodo}", 0)
            if peso_parcial > 0:
                datos_parciales.append(
                    {
                        "periodo": periodo,
                        "nota": notas_parciales[periodo],
                        "peso": peso_parcial,
                        "tiene_nota": notas_parciales[periodo] is not None
                        and notas_parciales[periodo] >= 0,
                    }
                )

            peso_continua = getattr(curso, f"peso_continua_{periodo}", 0)
            if peso_continua > 0:
                datos_continuas.append(
                    {
                        "periodo": periodo,
                        "nota": notas_continuas[periodo],
                        "peso": peso_continua,
                        "tiene_nota": notas_continuas[periodo] is not None
                        and notas_continuas[periodo] >= 0,
                    }
                )

        notas_por_curso.append(
            {
                "curso": curso,
                "nombre_curso": curso.nombre,
                "codigo_curso": curso.codigo,
                "turno": turno,
                "datos_parciales": datos_parciales,
                "datos_continuas": datos_continuas,
                "nota_sustitutorio": nota_sustitutorio,
                "tiene_sustitutorio": nota_sustitutorio is not None
                and nota_sustitutorio >= 0,
                "promedio_final": round(promedio_final, 2)
                if promedio_final is not None and promedio_final >= 0
                else None,
                "estado": estado_curso,
                "progreso": round(progreso_curso, 1),
                "total_notas_registradas": total_notas_curso,
                "total_notas_esperadas": total_notas_esperadas,
                "notas_totales": list(notas_curso),
                "configuracion_curso": {
                    "tiene_continua": any(
                        [
                            curso.peso_continua_1 > 0,
                            curso.peso_continua_2 > 0,
                            curso.peso_continua_3 > 0,
                        ]
                    ),
                    "tiene_parciales": any(
                        [
                            curso.peso_parcial_1 > 0,
                            curso.peso_parcial_2 > 0,
                            curso.peso_parcial_3 > 0,
                        ]
                    ),
                    "tiene_sustitutorio": (
                        curso.peso_parcial_1 > 0 and curso.peso_parcial_2 > 0
                    ),
                    "pesos_continua": {
                        1: curso.peso_continua_1,
                        2: curso.peso_continua_2,
                        3: curso.peso_continua_3,
                    },
                    "pesos_parcial": {
                        1: curso.peso_parcial_1,
                        2: curso.peso_parcial_2,
                        3: curso.peso_parcial_3,
                    },
                },
            }
        )

    # Calcular promedio general
    if cursos_con_promedio > 0:
        estadisticas_generales["promedio_general"] = round(
            suma_promedios / cursos_con_promedio, 2
        )

    # Calcular progreso general
    if total_notas_esperadas_global > 0:
        estadisticas_generales["progreso_general"] = round(
            (total_notas_registradas_global / total_notas_esperadas_global) * 100, 1
        )
    else:
        estadisticas_generales["progreso_general"] = 0

    # Calcular porcentajes para gráficos
    if estadisticas_generales["total_cursos"] > 0:
        estadisticas_generales["porcentaje_aprobados"] = round(
            (
                estadisticas_generales["cursos_aprobados"]
                / estadisticas_generales["total_cursos"]
            )
            * 100,
            1,
        )
        estadisticas_generales["porcentaje_desaprobados"] = round(
            (
                estadisticas_generales["cursos_desaprobados"]
                / estadisticas_generales["total_cursos"]
            )
            * 100,
            1,
        )
        estadisticas_generales["porcentaje_en_proceso"] = round(
            (
                estadisticas_generales["cursos_en_proceso"]
                / estadisticas_generales["total_cursos"]
            )
            * 100,
            1,
        )
    else:
        estadisticas_generales.update(
            {
                "porcentaje_aprobados": 0,
                "porcentaje_desaprobados": 0,
                "porcentaje_en_proceso": 0,
            }
        )

    # Ordenar cursos por estado y promedio
    notas_por_curso.sort(
        key=lambda x: (
            0 if x["estado"] == "aprobado" else 1 if x["estado"] == "en_proceso" else 2,
            -x["promedio_final"] if x["promedio_final"] else 0,
        )
    )

    context = {
        "alumno": alumno,
        "notas_por_curso": notas_por_curso,
        "estadisticas": estadisticas_generales,
        "semestre_actual": alumno.calcular_semestre() or "No asignado",
        "nombre": nombre,
        "rol": rol,
    }

    return render(request, "siscad/alumno/visualizar_notas.html", context)


def descargar_libreta_excel(alumno):
    """
    Genera y descarga un Excel con la libreta de notas del alumno
    """
    try:
        import pandas as pd
        import io
        from django.http import HttpResponse
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from django.utils import timezone

        # Obtener todas las matrículas del alumno
        matriculas = MatriculaCurso.objects.filter(alumno=alumno).select_related(
            "curso"
        )

        # Crear datos para el Excel
        data = []

        for matricula in matriculas:
            curso = matricula.curso
            notas_curso = Nota.objects.filter(alumno=alumno, curso=curso).order_by(
                "tipo", "periodo"
            )

            # Organizar notas
            notas_parciales = {1: "", 2: "", 3: ""}
            notas_continuas = {1: "", 2: "", 3: ""}
            nota_sustitutorio = ""

            for nota in notas_curso:
                if nota.tipo == "P" and nota.periodo in [1, 2, 3]:
                    if nota.valor is not None and nota.valor >= 0:
                        notas_parciales[nota.periodo] = nota.valor
                elif nota.tipo == "C" and nota.periodo in [1, 2, 3]:
                    if nota.valor is not None and nota.valor >= 0:
                        notas_continuas[nota.periodo] = nota.valor
                elif nota.tipo == "S":
                    if nota.valor is not None and nota.valor >= 0:
                        nota_sustitutorio = nota.valor

            # Calcular nota final
            nota_final = calcular_nota_final_alumno(alumno, curso)

            # Determinar estado
            estado = "En proceso"
            if nota_final is not None and nota_final >= 0:
                estado = "Aprobado" if nota_final >= 10.5 else "Desaprobado"

            data.append(
                {
                    "Código Curso": curso.codigo,
                    "Nombre Curso": curso.nombre,
                    "Turno": matricula.get_turno_display(),
                    "Continua 1": notas_continuas[1],
                    "Continua 2": notas_continuas[2],
                    "Continua 3": notas_continuas[3],
                    "Parcial 1": notas_parciales[1],
                    "Parcial 2": notas_parciales[2],
                    "Parcial 3": notas_parciales[3],
                    "Sustitutorio": nota_sustitutorio,
                    "Nota Final": nota_final
                    if nota_final is not None and nota_final >= 0
                    else "",
                    "Estado": estado,
                }
            )

        # Crear DataFrame
        df = pd.DataFrame(data)

        # Crear output
        output = io.BytesIO()

        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            # Hoja principal con datos
            df.to_excel(writer, sheet_name="Libreta de Notas", index=False)

            # Obtener la hoja para aplicar formato
            workbook = writer.book
            worksheet = writer.sheets["Libreta de Notas"]

            # Aplicar formato profesional
            header_font = Font(bold=True, color="FFFFFF", size=12)
            header_fill = PatternFill(
                start_color="366092", end_color="366092", fill_type="solid"
            )
            header_alignment = Alignment(horizontal="center", vertical="center")
            border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )

            # Aplicar estilo a los encabezados
            for cell in worksheet[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = border

            # Aplicar bordes a todas las celdas
            for row in worksheet.iter_rows(
                min_row=2, max_row=worksheet.max_row, min_col=1, max_col=len(df.columns)
            ):
                for cell in row:
                    cell.border = border

            # Ajustar ancho de columnas
            column_widths = {
                "A": 12,
                "B": 30,
                "C": 10,
                "D": 10,
                "E": 10,
                "F": 10,
                "G": 10,
                "H": 10,
                "I": 10,
                "J": 12,
                "K": 12,
                "L": 12,
            }

            for col, width in column_widths.items():
                worksheet.column_dimensions[col].width = width

            # Aplicar formato condicional para estados
            for row in range(2, worksheet.max_row + 1):
                estado_cell = worksheet[f"L{row}"]
                if estado_cell.value == "Aprobado":
                    estado_cell.fill = PatternFill(
                        start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"
                    )
                    estado_cell.font = Font(color="006100", bold=True)
                elif estado_cell.value == "Desaprobado":
                    estado_cell.fill = PatternFill(
                        start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"
                    )
                    estado_cell.font = Font(color="9C0006", bold=True)
                else:
                    estado_cell.fill = PatternFill(
                        start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"
                    )
                    estado_cell.font = Font(color="9C6500", bold=True)

            # Hoja de resumen
            resumen_data = [
                ["ALUMNO", f"{alumno.nombre}"],
                ["DNI", f"{alumno.dni}"],
                ["SEMESTRE", f"{alumno.calcular_semestre() or 'No asignado'}"],
                ["FECHA DE GENERACIÓN", f"{timezone.now().strftime('%Y-%m-%d %H:%M')}"],
                ["", ""],
                ["RESUMEN ACADÉMICO", ""],
                ["Total de Cursos", len(data)],
                [
                    "Cursos Aprobados",
                    len([d for d in data if d["Estado"] == "Aprobado"]),
                ],
                [
                    "Cursos Desaprobados",
                    len([d for d in data if d["Estado"] == "Desaprobado"]),
                ],
                [
                    "Cursos en Proceso",
                    len([d for d in data if d["Estado"] == "En proceso"]),
                ],
                ["Promedio General", f"{calcular_promedio_general(alumno):.2f}"],
            ]

            df_resumen = pd.DataFrame(resumen_data, columns=["Concepto", "Valor"])
            df_resumen.to_excel(writer, sheet_name="Resumen", index=False)

            # Formato hoja de resumen
            worksheet_resumen = writer.sheets["Resumen"]
            worksheet_resumen.column_dimensions["A"].width = 25
            worksheet_resumen.column_dimensions["B"].width = 20

            # Aplicar formato a la hoja de resumen
            for cell in worksheet_resumen[1]:
                cell.font = Font(bold=True, color="FFFFFF", size=12)
                cell.fill = PatternFill(
                    start_color="28a745", end_color="28a745", fill_type="solid"
                )
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = border

            for row in range(2, worksheet_resumen.max_row + 1):
                for col in ["A", "B"]:
                    cell = worksheet_resumen[f"{col}{row}"]
                    cell.border = border
                    if row >= 7:  # Filas del resumen académico
                        cell.font = Font(bold=True)

        output.seek(0)

        # Crear respuesta HTTP
        response = HttpResponse(
            output.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = (
            f'attachment; filename="libreta_notas_{alumno.dni}_{timezone.now().strftime("%Y%m%d")}.xlsx"'
        )

        return response

    except Exception as e:
        # En un entorno real, deberías manejar este error adecuadamente
        from django.contrib import messages

        return redirect("visualizar_notas")


def calcular_promedio_general(alumno):
    """
    Calcula el promedio general del alumno
    """
    matriculas = MatriculaCurso.objects.filter(alumno=alumno).select_related("curso")
    suma_promedios = 0
    cursos_con_promedio = 0

    for matricula in matriculas:
        curso = matricula.curso
        nota_final = calcular_nota_final_alumno(alumno, curso)
        if nota_final is not None and nota_final >= 0:
            suma_promedios += nota_final
            cursos_con_promedio += 1

    return suma_promedios / cursos_con_promedio if cursos_con_promedio > 0 else 0


# Funciones auxiliares actualizadas para trabajar con -1
def calcular_nota_final_alumno(alumno, curso):
    """
    Calcula la nota final para un alumno en un curso específico - VERSIÓN CON -1
    """
    notas = Nota.objects.filter(alumno=alumno, curso=curso)

    # Solo considerar notas válidas (>= 0)
    continuas = {
        n.periodo: n
        for n in notas
        if n.tipo == "C" and n.valor is not None and n.valor >= 0
    }
    parciales = {
        n.periodo: n
        for n in notas
        if n.tipo == "P" and n.valor is not None and n.valor >= 0
    }
    sustitutorio = next(
        (n for n in notas if n.tipo == "S" and n.valor is not None and n.valor >= 0),
        None,
    )

    # Aplicar sustitutorio si existe
    if (
        sustitutorio
        and sustitutorio.valor is not None
        and sustitutorio.valor >= 0
        and 1 in parciales
        and 2 in parciales
    ):
        if parciales[1].valor <= parciales[2].valor:
            nota_original_p1 = parciales[1].valor
            parciales[1].valor = sustitutorio.valor
        else:
            nota_original_p2 = parciales[2].valor
            parciales[2].valor = sustitutorio.valor

    # Calcular promedio ponderado
    total_puntos = 0
    total_pesos = 0

    # Sumar continuas
    for periodo in [1, 2, 3]:
        if periodo in continuas:
            peso = getattr(curso, f"peso_continua_{periodo}", 0)
            if peso > 0:
                total_puntos += continuas[periodo].valor * peso
                total_pesos += peso

    # Sumar parciales
    for periodo in [1, 2, 3]:
        if periodo in parciales:
            peso = getattr(curso, f"peso_parcial_{periodo}", 0)
            if peso > 0:
                total_puntos += parciales[periodo].valor * peso
                total_pesos += peso

    if total_pesos == 0:
        return None

    nota_final = total_puntos / total_pesos

    # Restaurar notas originales si se aplicó sustitutorio
    if (
        sustitutorio
        and sustitutorio.valor is not None
        and sustitutorio.valor >= 0
        and 1 in parciales
        and 2 in parciales
    ):
        if "nota_original_p1" in locals():
            parciales[1].valor = nota_original_p1
        elif "nota_original_p2" in locals():
            parciales[2].valor = nota_original_p2

    return round(nota_final, 2)


def determinar_estado_curso(promedio_final, curso, notas_parciales, notas_continuas):
    """
    Determina el estado del curso (aprobado, desaprobado, en proceso) - VERSIÓN CON -1
    """
    if promedio_final is None or promedio_final < 0:
        return "en_proceso"

    # Verificar si todas las notas necesarias están registradas
    notas_faltantes = False

    # Verificar parciales
    for periodo in [1, 2, 3]:
        peso_parcial = getattr(curso, f"peso_parcial_{periodo}", 0)
        if peso_parcial > 0 and (
            notas_parciales[periodo] is None or notas_parciales[periodo] < 0
        ):
            notas_faltantes = True

    # Verificar continuas
    for periodo in [1, 2, 3]:
        peso_continua = getattr(curso, f"peso_continua_{periodo}", 0)
        if peso_continua > 0 and (
            notas_continuas[periodo] is None or notas_continuas[periodo] < 0
        ):
            notas_faltantes = True

    if notas_faltantes:
        return "en_proceso"

    # Determinar aprobación basado en el promedio final
    if promedio_final >= 10.5:  # Umbral de aprobación
        return "aprobado"
    else:
        return "desaprobado"


def calcular_total_notas_esperadas(curso):
    """
    Calcula el total de notas esperadas para un curso según su configuración
    """
    total = 0

    # Contar parciales esperados
    for periodo in [1, 2, 3]:
        if getattr(curso, f"peso_parcial_{periodo}", 0) > 0:
            total += 1

    # Contar continuas esperadas
    for periodo in [1, 2, 3]:
        if getattr(curso, f"peso_continua_{periodo}", 0) > 0:
            total += 1

    # Agregar sustitutorio si aplica
    if curso.peso_parcial_1 > 0 and curso.peso_parcial_2 > 0:
        total += 1

    return total


def visualizar_horario_alumno(request):
    # 1. Obtener el alumno logueado usando el email de la sesión
    if "email" not in request.session:
        return redirect("login")

    email = request.session["email"]

    try:
        alumno = Alumno.objects.get(email=email)
    except Alumno.DoesNotExist:
        return redirect("login")
    except Alumno.MultipleObjectsReturned:
        alumno = Alumno.objects.filter(email=email).first()

    # 2. Obtener todas sus matrículas de curso
    matriculas_curso = MatriculaCurso.objects.filter(alumno=alumno).select_related(
        "curso"
    )
    print(
        f"DEBUG: Matrículas de curso del alumno: {[(m.curso.nombre, m.turno) for m in matriculas_curso]}"
    )

    # 3. Obtener todas sus matrículas de laboratorio
    matriculas_lab = MatriculaLaboratorio.objects.filter(alumno=alumno).select_related(
        "grupo_laboratorio__grupo_teoria__curso", "grupo_laboratorio__profesor"
    )
    print(
        f"DEBUG: Matrículas de laboratorio del alumno: {[f'{ml.grupo_laboratorio.grupo_teoria.curso.nombre} - Lab {ml.grupo_laboratorio.grupo}' for ml in matriculas_lab]}"
    )

    # 4. Obtener todos los cursos inscritos y sus turnos (para teoría y práctica)
    cursos_turnos = {m.curso_id: m.turno for m in matriculas_curso}
    cursos_ids = list(cursos_turnos.keys())

    # 5. Obtener IDs de grupos de laboratorio matriculados (para laboratorios)
    grupos_lab_ids = [ml.grupo_laboratorio_id for ml in matriculas_lab]

    if not cursos_ids and not grupos_lab_ids:
        cursos_ids = []
        grupos_lab_ids = []

    # 6. Buscar horarios relacionados - AHORA INCLUYENDO LABORATORIOS
    horarios = (
        Hora.objects.select_related(
            "aula",
            "grupo_teoria__curso",
            "grupo_practica__grupo_teoria__curso",
            "grupo_laboratorio__grupo_teoria__curso",
            "grupo_laboratorio__profesor",
        )
        .filter(
            # Coincidencia con teoría del mismo turno
            Q(
                grupo_teoria__curso_id__in=cursos_ids,
                grupo_teoria__turno__in=[turno for turno in cursos_turnos.values()],
            )
            |
            # Coincidencia con práctica del mismo turno
            Q(
                grupo_practica__grupo_teoria__curso_id__in=cursos_ids,
                grupo_practica__turno__in=[turno for turno in cursos_turnos.values()],
            )
            |
            # Coincidencia con laboratorios matriculados
            Q(grupo_laboratorio_id__in=grupos_lab_ids)
        )
        .order_by("dia", "hora_inicio")
    )

    print(
        f"DEBUG: Horarios encontrados (teoría + práctica + laboratorio): {horarios.count()}"
    )

    # Debug detallado de horarios encontrados
    for h in horarios:
        if h.grupo_teoria:
            tipo = "Teoría"
            curso_nombre = h.grupo_teoria.curso.nombre
            turno_info = f"T-{h.grupo_teoria.turno}"
        elif h.grupo_practica:
            tipo = "Práctica"
            curso_nombre = h.grupo_practica.grupo_teoria.curso.nombre
            turno_info = f"P-{h.grupo_practica.turno}"
        elif h.grupo_laboratorio:
            tipo = "Laboratorio"
            curso_nombre = h.grupo_laboratorio.grupo_teoria.curso.nombre
            turno_info = f"L-{h.grupo_laboratorio.grupo}"
        else:
            tipo = "Otro"
            curso_nombre = "Desconocido"
            turno_info = ""

        print(
            f"DEBUG Horario: {h.dia} {h.hora_inicio}-{h.hora_fin} - {tipo}: {curso_nombre} {turno_info} - Aula: {h.aula.nombre if h.aula else 'Sin aula'}"
        )

    # 7. Construir estructura para mostrar
    dias_lista = [
        ("L", "Lunes"),
        ("M", "Martes"),
        ("X", "Miércoles"),
        ("J", "Jueves"),
        ("V", "Viernes"),
    ]

    # Crear estructura de datos
    tabla_horarios = {}
    for h in horarios:
        bloque = f"{h.hora_inicio.strftime('%H:%M')} - {h.hora_fin.strftime('%H:%M')}"
        dia = h.dia

        # Determinar si el horario debe mostrarse (filtrado por turno para teoría/práctica)
        mostrar_horario = False
        info_curso = ""

        if h.grupo_teoria:
            curso_id = h.grupo_teoria.curso_id
            turno_horario = h.grupo_teoria.turno
            turno_alumno = cursos_turnos.get(curso_id)

            # Mostrar solo si el turno coincide
            if turno_horario == turno_alumno:
                mostrar_horario = True
                curso = h.grupo_teoria.curso.nombre
                grupo = f"T-{turno_horario}"
                aula_info = f" ({h.aula.nombre})" if h.aula else ""
                info_curso = f"{curso} {grupo}{aula_info}"

        elif h.grupo_practica:
            curso_id = h.grupo_practica.grupo_teoria.curso_id
            turno_horario = h.grupo_practica.turno
            turno_alumno = cursos_turnos.get(curso_id)

            # Mostrar solo si el turno coincide
            if turno_horario == turno_alumno:
                mostrar_horario = True
                curso = h.grupo_practica.grupo_teoria.curso.nombre
                grupo = f"P-{turno_horario}"
                aula_info = f" ({h.aula.nombre})" if h.aula else ""
                info_curso = f"{curso} {grupo}{aula_info}"

        elif h.grupo_laboratorio:
            # Para laboratorios, mostrar siempre (ya están filtrados por matrícula)
            mostrar_horario = True
            curso = h.grupo_laboratorio.grupo_teoria.curso.nombre
            grupo = f"L-{h.grupo_laboratorio.grupo}"
            aula_info = f" ({h.aula.nombre})" if h.aula else ""
            profesor_info = (
                f" - {h.grupo_laboratorio.profesor.nombre}"
                if h.grupo_laboratorio.profesor
                else ""
            )
            info_curso = f"{curso} {grupo}{aula_info}{profesor_info}"

        # Si el horario debe mostrarse, agregarlo a la tabla
        if mostrar_horario and info_curso:
            # Usar "Horario General" como clave única para simplificar
            aula_key = "Horario General"

            if aula_key not in tabla_horarios:
                tabla_horarios[aula_key] = {}

            if bloque not in tabla_horarios[aula_key]:
                tabla_horarios[aula_key][bloque] = {d: "" for d, _ in dias_lista}

            tabla_horarios[aula_key][bloque][dia] = info_curso

    # Ordenar por hora
    for aula in tabla_horarios:
        tabla_horarios[aula] = dict(
            sorted(tabla_horarios[aula].items(), key=lambda x: x[0])
        )

    # Crear la estructura final
    context = {
        "alumno": alumno,
        "dias_lista": dias_lista,
        "tabla_horarios": {
            aula: [
                {
                    "bloque": bloque,
                    "dias": [(dia, dias_data[dia]) for dia, _ in dias_lista],
                }
                for bloque, dias_data in bloques.items()
            ]
            for aula, bloques in tabla_horarios.items()
        },
        "tiene_laboratorios": len(grupos_lab_ids) > 0,
        "total_laboratorios": len(grupos_lab_ids),
    }

    return render(request, "siscad/alumno/visualizar_horario_alumno.html", context)


def matricula_laboratorio(request):
    # 1. Obtener el alumno logueado
    if "email" not in request.session:
        return redirect("login")

    email = request.session["email"]

    try:
        alumno = Alumno.objects.get(email=email)
    except Alumno.DoesNotExist:
        messages.error(request, "Alumno no encontrado.")
        return redirect("login")
    except Alumno.MultipleObjectsReturned:
        alumno = Alumno.objects.filter(email=email).first()

    # 2. Obtener los cursos matriculados
    matriculas_curso = MatriculaCurso.objects.filter(
        alumno_id=alumno.id
    ).select_related("curso")

    if not matriculas_curso:
        messages.warning(request, "No tienes cursos matriculados.")
        return redirect("inicio_alumno")

    # 3. Procesar matrícula de laboratorio
    if request.method == "POST":
        grupo_lab_id = request.POST.get("grupo_laboratorio")

        if not grupo_lab_id:
            messages.error(request, "Debes seleccionar un grupo de laboratorio.")
            return redirect("matricula_laboratorio")

        try:
            with transaction.atomic():
                grupo_lab = GrupoLaboratorio.objects.select_related(
                    "grupo_teoria__curso", "profesor"
                ).get(id=grupo_lab_id)

                # Verificar que el alumno esté matriculado en el curso correspondiente
                matricula_curso = matriculas_curso.filter(
                    curso_id=grupo_lab.grupo_teoria.curso_id
                ).first()

                if not matricula_curso:
                    messages.error(request, "No estás matriculado en este curso.")
                    return redirect("matricula_laboratorio")

                # Verificar si ya está matriculado en un laboratorio de ese curso
                matricula_existente = MatriculaLaboratorio.objects.filter(
                    alumno_id=alumno.id,
                    grupo_laboratorio__grupo_teoria__curso_id=grupo_lab.grupo_teoria.curso_id,
                ).exists()

                if matricula_existente:
                    messages.error(
                        request,
                        f"Ya estás matriculado en un laboratorio de {grupo_lab.grupo_teoria.curso.nombre}.",
                    )
                    return redirect("matricula_laboratorio")

                # Verificar cupos
                if grupo_lab.cupos <= 0:
                    messages.error(
                        request,
                        "No hay cupos disponibles en este grupo de laboratorio.",
                    )
                    return redirect("matricula_laboratorio")

                # Crear matrícula
                matricula_lab = MatriculaLaboratorio(
                    alumno_id=alumno.id, grupo_laboratorio_id=grupo_lab.id
                )
                matricula_lab.save()

                # Actualizar cupos
                grupo_lab.cupos -= 1
                grupo_lab.save()

                generar_asistencias_laboratorio(alumno, grupo_lab)

                messages.success(
                    request,
                    f"Te has matriculado exitosamente en el laboratorio de {grupo_lab.grupo_teoria.curso.nombre} y se generaron tus asistencias.",
                )
                return redirect("matricula_laboratorio")

        except GrupoLaboratorio.DoesNotExist:
            messages.error(request, "El grupo de laboratorio seleccionado no existe.")
        except Exception as e:
            messages.error(request, f"Error al realizar la matrícula: {str(e)}")

    # 4. Obtener laboratorios disponibles
    laboratorios_disponibles = []

    for matricula in matriculas_curso:
        curso = matricula.curso
        turno_alumno = matricula.turno

        grupos_lab = (
            GrupoLaboratorio.objects.select_related("grupo_teoria__curso", "profesor")
            .filter(
                grupo_teoria__curso_id=curso.id,
                cupos__gt=0,
            )
            .exclude(matriculas_laboratorio__alumno_id=alumno.id)
        )

        grupos_compatibles = [
            g
            for g in grupos_lab
            if g.grupo == turno_alumno or g.grupo not in ["A", "B", "C"]
        ]

        if grupos_compatibles:
            laboratorios_disponibles.append(
                {
                    "curso": curso,
                    "turno_alumno": turno_alumno,
                    "grupos": grupos_compatibles,
                    "ya_matriculado": MatriculaLaboratorio.objects.filter(
                        alumno_id=alumno.id,
                        grupo_laboratorio__grupo_teoria__curso_id=curso.id,
                    ).exists(),
                }
            )

    # 5. Matrículas actuales
    matriculas_lab_actuales = MatriculaLaboratorio.objects.filter(
        alumno_id=alumno.id
    ).select_related(
        "grupo_laboratorio__grupo_teoria__curso", "grupo_laboratorio__profesor"
    )

    context = {
        "alumno": alumno,
        "laboratorios_disponibles": laboratorios_disponibles,
        "matriculas_lab_actuales": matriculas_lab_actuales,
    }

    return render(request, "siscad/alumno/matricula_laboratorio.html", context)


def generar_asistencias_laboratorio(alumno, grupo_lab):
    """
    Genera asistencias para el laboratorio matriculado por el alumno,
    desde el 2 de septiembre 2025 hasta el 25 de diciembre 2025.
    Si el laboratorio tiene horas consecutivas (bloques), se genera una sola asistencia por bloque.
    """
    print(
        f"🔧 Generando asistencias de laboratorio para {alumno.nombre} - {grupo_lab.grupo_teoria.curso.nombre}"
    )

    # Rango de fechas
    fecha_inicio = date(2025, 9, 2)
    fecha_fin = date(2025, 12, 25)
    fecha_hoy = date.today()

    asistencias_creadas = 0

    # Obtener los horarios del grupo de laboratorio
    horarios_lab = (
        Hora.objects.filter(grupo_laboratorio=grupo_lab)
        .select_related("grupo_laboratorio__grupo_teoria__curso")
        .order_by("dia", "hora_inicio")
    )

    if not horarios_lab.exists():
        print(f"   ⚠️ No se encontraron horarios para el laboratorio {grupo_lab.id}")
        return 0

    # Mapear días
    dias_map = {
        "Monday": "L",
        "Tuesday": "M",
        "Wednesday": "X",
        "Thursday": "J",
        "Friday": "V",
    }

    # Recorrer días desde inicio a fin
    fecha_actual = fecha_inicio
    while fecha_actual <= fecha_fin:
        if fecha_actual.weekday() < 5:  # Solo lunes a viernes
            dia_codigo = dias_map.get(fecha_actual.strftime("%A"))

            # Horarios del laboratorio en ese día
            horarios_dia = [h for h in horarios_lab if h.dia == dia_codigo]
            if horarios_dia:
                # Agrupar bloques consecutivos
                horarios_dia.sort(key=lambda h: h.hora_inicio)
                bloques = []
                bloque_actual = [horarios_dia[0]]

                for i in range(1, len(horarios_dia)):
                    anterior = bloque_actual[-1]
                    actual = horarios_dia[i]
                    if actual.hora_inicio == anterior.hora_fin:
                        bloque_actual.append(actual)
                    else:
                        bloques.append(bloque_actual)
                        bloque_actual = [actual]
                bloques.append(bloque_actual)

                estado = "P" if fecha_actual <= fecha_hoy else "F"

                for bloque in bloques:
                    try:
                        hora_inicio = bloque[0].hora_inicio
                        hora_fin = bloque[-1].hora_fin
                        hora_referencia = bloque[0]

                        # Evitar duplicados
                        asistencia_existente = AsistenciaAlumno.objects.filter(
                            alumno=alumno,
                            fecha=fecha_actual,
                            hora__grupo_laboratorio=grupo_lab,
                        ).exists()

                        if asistencia_existente:
                            continue

                        # Crear asistencia
                        asistencia = AsistenciaAlumno(
                            alumno=alumno,
                            fecha=fecha_actual,
                            estado=estado,
                            hora=hora_referencia,
                        )
                        asistencia.save()
                        asistencias_creadas += 1

                        estado_display = "PRESENTE" if estado == "P" else "FALTA"
                        print(
                            f"   {fecha_actual} - {grupo_lab.grupo_teoria.curso.nombre} (Lab) - {estado_display}"
                        )

                    except Exception as e:
                        print(f"   ❌ Error generando asistencia {fecha_actual}: {e}")
                        continue

        fecha_actual += timedelta(days=1)

    print(f"✅ Total asistencias generadas: {asistencias_creadas}")
    return asistencias_creadas


def visualizar_asistencias_alumno(request):
    # 1. Obtener el alumno logueado
    if "email" not in request.session:
        return redirect("login")

    email = request.session["email"]

    try:
        alumno = Alumno.objects.get(email=email)
    except Alumno.DoesNotExist:
        return redirect("login")
    except Alumno.MultipleObjectsReturned:
        alumno = Alumno.objects.filter(email=email).first()

    # 2. Obtener cursos matriculados (tanto de curso como de laboratorio)
    cursos_matriculados = obtener_cursos_alumno(alumno)

    # 3. Procesar filtro por curso si se envió
    curso_seleccionado_id = request.GET.get("curso_id")
    curso_seleccionado = None
    asistencias_curso = []
    estadisticas_curso = {}

    if curso_seleccionado_id:
        try:
            curso_seleccionado = Curso.objects.get(id=curso_seleccionado_id)
            asistencias_curso = obtener_asistencias_curso(alumno, curso_seleccionado)
            estadisticas_curso = calcular_estadisticas_asistencias(asistencias_curso)
        except Curso.DoesNotExist:
            pass

    # 4. Preparar contexto
    context = {
        "alumno": alumno,
        "cursos_matriculados": cursos_matriculados,
        "curso_seleccionado": curso_seleccionado,
        "asistencias_curso": asistencias_curso,
        "estadisticas_curso": estadisticas_curso,
        "hoy": datetime.now().date(),  # Fecha actual corregida para 2025
        "anio_actual": 2025,
    }

    return render(request, "siscad/alumno/visualizar_asistencias_alumno.html", context)


def obtener_cursos_alumno(alumno):
    """
    Obtiene todos los cursos en los que el alumno está matriculado
    (tanto de MatriculaCurso como de MatriculaLaboratorio)
    """
    cursos = []

    # Cursos de teoría/práctica
    matriculas_curso = MatriculaCurso.objects.filter(alumno=alumno).select_related(
        "curso"
    )
    for matricula in matriculas_curso:
        cursos.append(
            {
                "id": matricula.curso.id,
                "nombre": matricula.curso.nombre,
                "tipo": "Curso",
                "turno": matricula.turno,
                "matricula_id": matricula.id,
            }
        )

    # Cursos de laboratorio (sin duplicados con los de teoría)
    matriculas_lab = MatriculaLaboratorio.objects.filter(alumno=alumno).select_related(
        "grupo_laboratorio__grupo_teoria__curso"
    )

    for matricula in matriculas_lab:
        curso = matricula.grupo_laboratorio.grupo_teoria.curso
        # Solo agregar si no está ya en la lista
        if not any(c["id"] == curso.id for c in cursos):
            cursos.append(
                {
                    "id": curso.id,
                    "nombre": curso.nombre,
                    "tipo": "Laboratorio",
                    "grupo_lab": matricula.grupo_laboratorio.grupo,
                    "matricula_id": matricula.id,
                }
            )

    return cursos


def obtener_asistencias_curso(alumno, curso):
    """
    Obtiene todas las asistencias del alumno para un curso específico
    """
    # Obtener asistencias relacionadas con este curso
    asistencias = (
        AsistenciaAlumno.objects.filter(alumno=alumno)
        .filter(
            Q(hora__grupo_teoria__curso=curso)
            | Q(hora__grupo_practica__grupo_teoria__curso=curso)
            | Q(hora__grupo_laboratorio__grupo_teoria__curso=curso)
        )
        .select_related(
            "hora__grupo_teoria__curso",
            "hora__grupo_practica__grupo_teoria__curso",
            "hora__grupo_laboratorio__grupo_teoria__curso",
        )
        .order_by("-fecha", "hora__hora_inicio")
    )

    # Formatear los datos para el template
    asistencias_formateadas = []
    for asistencia in asistencias:
        # Determinar tipo de clase y información
        tipo_clase = "Desconocido"
        detalle_clase = ""

        if asistencia.hora.grupo_teoria:
            tipo_clase = "Teoría"
            detalle_clase = f"T-{asistencia.hora.grupo_teoria.turno}"
        elif asistencia.hora.grupo_practica:
            tipo_clase = "Práctica"
            detalle_clase = f"P-{asistencia.hora.grupo_practica.turno}"
        elif asistencia.hora.grupo_laboratorio:
            tipo_clase = "Laboratorio"
            detalle_clase = f"L-{asistencia.hora.grupo_laboratorio.grupo}"

        asistencias_formateadas.append(
            {
                "fecha": asistencia.fecha,
                "dia_semana": asistencia.fecha.strftime("%A"),
                "hora_inicio": asistencia.hora.hora_inicio.strftime("%H:%M"),
                "hora_fin": asistencia.hora.hora_fin.strftime("%H:%M"),
                "tipo_clase": tipo_clase,
                "detalle_clase": detalle_clase,
                "estado": asistencia.estado,
                "estado_display": asistencia.get_estado_display(),
                "aula": asistencia.hora.aula.nombre
                if asistencia.hora.aula
                else "Sin aula",
                "es_pasado": asistencia.fecha < datetime.now().date(),
            }
        )

    return asistencias_formateadas


def calcular_estadisticas_asistencias(asistencias):
    """
    Calcula estadísticas de asistencias para un curso
    """
    if not asistencias:
        return {}

    total = len(asistencias)
    presentes = len([a for a in asistencias if a["estado"] == "P"])
    faltas = len([a for a in asistencias if a["estado"] == "F"])

    # Separar por tipo de clase
    teorias = [a for a in asistencias if a["tipo_clase"] == "Teoría"]
    practicas = [a for a in asistencias if a["tipo_clase"] == "Práctica"]
    laboratorios = [a for a in asistencias if a["tipo_clase"] == "Laboratorio"]

    # Calcular porcentajes
    porcentaje_asistencia = (presentes / total * 100) if total > 0 else 0

    return {
        "total": total,
        "presentes": presentes,
        "faltas": faltas,
        "porcentaje_asistencia": round(porcentaje_asistencia, 1),
        "teorias_total": len(teorias),
        "teorias_presentes": len([t for t in teorias if t["estado"] == "P"]),
        "practicas_total": len(practicas),
        "practicas_presentes": len([p for p in practicas if p["estado"] == "P"]),
        "laboratorios_total": len(laboratorios),
        "laboratorios_presentes": len([l for l in laboratorios if l["estado"] == "P"]),
    }


# =======================Vista de Profesor====================================================
def inicio_profesor(request):
    rol = request.session.get("rol")

    if rol != "Profesor":
        request.session["rol"] = "Ninguno"
        return redirect("login")

    nombre = request.session.get("nombre")
    return render(request, "siscad/profesor/menu.html", {"nombre": nombre, "rol": rol})


def visualizar_horario_profesor(request):
    if "email" not in request.session:
        return redirect("login")

    profesor = get_object_or_404(Profesor, email=request.session["email"])

    hoy = date.today()
    dia_semana = hoy.weekday()  # lunes=0 ... domingo=6
    if dia_semana >= 5:  # si es sábado o domingo → próxima semana
        inicio_semana = hoy + timedelta(days=(7 - dia_semana))
    else:
        inicio_semana = hoy - timedelta(days=dia_semana)

    fin_semana = inicio_semana + timedelta(days=4)  # lunes a viernes

    reservas_semana = Reserva.objects.filter(
        profesor=profesor, fecha__range=(inicio_semana, fin_semana)
    ).select_related("aula")

    horarios = (
        Hora.objects.select_related(
            "aula",
            "grupo_teoria__curso",
            "grupo_practica__grupo_teoria__curso",
            "grupo_laboratorio__grupo_teoria__curso",
        )
        .filter(
            Q(grupo_teoria__profesor=profesor)
            | Q(grupo_practica__profesor=profesor)
            | Q(grupo_laboratorio__profesor=profesor)
        )
        .order_by("dia", "hora_inicio")
    )

    dias_lista = [
        ("L", "Lunes"),
        ("M", "Martes"),
        ("X", "Miércoles"),
        ("J", "Jueves"),
        ("V", "Viernes"),
    ]

    tabla_horarios = {}
    for h in horarios:
        bloque = f"{h.hora_inicio.strftime('%H:%M')} - {h.hora_fin.strftime('%H:%M')}"
        dia = h.dia

        aula_key = "Horario del Profesor"
        if aula_key not in tabla_horarios:
            tabla_horarios[aula_key] = {}
        if bloque not in tabla_horarios[aula_key]:
            tabla_horarios[aula_key][bloque] = {d: "" for d, _ in dias_lista}

        if h.grupo_teoria:
            curso = h.grupo_teoria.curso.nombre
            grupo = f"T-{h.grupo_teoria.turno}"
            tipo_info = "Teoría"
        elif h.grupo_practica:
            curso = h.grupo_practica.grupo_teoria.curso.nombre
            grupo = f"P-{h.grupo_practica.turno}"
            tipo_info = "Práctica"
        elif h.grupo_laboratorio:
            curso = h.grupo_laboratorio.grupo_teoria.curso.nombre
            grupo = f"L-{h.grupo_laboratorio.grupo}"
            tipo_info = "Laboratorio"
        else:
            curso, grupo, tipo_info = "Sin asignar", "", ""

        aula_info = f" ({h.aula.nombre})" if h.aula else ""
        info_curso = f"{curso} {grupo} - {tipo_info}{aula_info}"

        tabla_horarios[aula_key][bloque][dia] = info_curso

    for r in reservas_semana:
        dia_letra = ["L", "M", "X", "J", "V"][r.fecha.weekday()]
        horas_reserva = Hora.objects.filter(reserva=r)

        for h in horas_reserva:
            bloque = (
                f"{h.hora_inicio.strftime('%H:%M')} - {h.hora_fin.strftime('%H:%M')}"
            )
            aula_key = "Horario del Profesor"
            if aula_key not in tabla_horarios:
                tabla_horarios[aula_key] = {}
            if bloque not in tabla_horarios[aula_key]:
                tabla_horarios[aula_key][bloque] = {d: "" for d, _ in dias_lista}

            tabla_horarios[aula_key][bloque][dia_letra] = (
                f"🟢 Reserva ({r.aula.nombre})"
            )

    for aula in tabla_horarios:
        tabla_horarios[aula] = dict(
            sorted(tabla_horarios[aula].items(), key=lambda x: x[0])
        )

    total_reservas_activas = Reserva.objects.filter(
        profesor=profesor, fecha__gte=hoy
    ).count()
    reservas_disponibles = max(0, profesor.cantidad_reservas - total_reservas_activas)

    estadisticas = {
        "total_horarios": horarios.count(),
        "total_teoria": horarios.filter(grupo_teoria__profesor=profesor).count(),
        "total_practica": horarios.filter(grupo_practica__profesor=profesor).count(),
        "total_laboratorio": horarios.filter(
            grupo_laboratorio__profesor=profesor
        ).count(),
        "cursos_unicos": len(
            set(
                [h.grupo_teoria.curso_id for h in horarios if h.grupo_teoria]
                + [
                    h.grupo_practica.grupo_teoria.curso_id
                    for h in horarios
                    if h.grupo_practica
                ]
                + [
                    h.grupo_laboratorio.grupo_teoria.curso_id
                    for h in horarios
                    if h.grupo_laboratorio
                ]
            )
        ),
        "reservas_disponibles": reservas_disponibles,
        "reservas_activas": total_reservas_activas,
        "semana": f"{inicio_semana.strftime('%d/%m/%Y')} - {fin_semana.strftime('%d/%m/%Y')}",
    }

    context = {
        "profesor": profesor,
        "dias_lista": dias_lista,
        "tabla_horarios": {
            aula: [
                {
                    "bloque": bloque,
                    "dias": [(dia, dias_data[dia]) for dia, _ in dias_lista],
                }
                for bloque, dias_data in bloques.items()
            ]
            for aula, bloques in tabla_horarios.items()
        },
        "estadisticas": estadisticas,
    }

    return render(request, "siscad/profesor/visualizar_horario_profesor.html", context)


def reservar_aula(request):
    # Verificar sesión
    if "email" not in request.session:
        return redirect("login")

    profesor = get_object_or_404(Profesor, email=request.session["email"])

    # 🔹 Eliminar reservas pasadas automáticamente
    Reserva.objects.filter(fecha__lt=date.today()).delete()

    # 🔹 Contar reservas activas (de hoy o futuras)
    reservas_activas_count = Reserva.objects.filter(
        profesor=profesor,
        fecha__gte=date.today(),
    ).count()

    # 🔹 Verificar si alcanzó el límite de reservas activas
    if reservas_activas_count >= profesor.cantidad_reservas:
        messages.error(request, "Ya has alcanzado el máximo de reservas permitidas.")
        return render(
            request,
            "siscad/profesor/reservar_aula.html",
            {
                "aulas": Aula.objects.all(),
                "horas": [],
                "reservas": Reserva.objects.filter(profesor=profesor),
                "fecha_seleccionada": None,
                "dia_letra": None,
            },
        )

    aulas = Aula.objects.all()
    horas_disponibles = []
    fecha_seleccionada = None
    dia_letra = None

    # -------------------- POST (reservar aula) --------------------
    if request.method == "POST":
        aula_id = request.POST.get("aula_id")
        fecha = request.POST.get("fecha")
        hora_inicio = request.POST.get("hora_inicio")

        if not (aula_id and fecha and hora_inicio):
            messages.error(request, "Por favor completa todos los campos.")
            return render(
                request,
                "siscad/profesor/reservar_aula.html",
                {
                    "aulas": aulas,
                    "horas": [],
                    "reservas": Reserva.objects.filter(profesor=profesor),
                    "fecha_seleccionada": None,
                    "dia_letra": None,
                },
            )

        aula = get_object_or_404(Aula, id=aula_id)
        fecha_seleccionada = datetime.strptime(fecha, "%Y-%m-%d").date()

        # Día de la semana (L, M, X, J, V)
        dia_semana = fecha_seleccionada.weekday()
        if dia_semana >= 5:
            messages.error(request, "Solo puedes reservar entre lunes y viernes.")
            return redirect("reservar_aula")

        letras_dia = ["L", "M", "X", "J", "V"]
        dia_letra = letras_dia[dia_semana]

        # Verificar que la hora esté libre
        hora_obj = get_object_or_404(
            Hora, aula=aula, dia=dia_letra, hora_inicio=hora_inicio, tipo__isnull=True
        )

        # Verificar conflictos con clases o reservas del profesor
        conflicto_profesor = Hora.objects.filter(
            Q(grupo_teoria__profesor=profesor)
            | Q(grupo_practica__profesor=profesor)
            | Q(grupo_laboratorio__profesor=profesor)
            | Q(reserva__profesor=profesor),
            dia=dia_letra,
            hora_inicio=hora_obj.hora_inicio,
        ).exists()

        if conflicto_profesor:
            messages.error(request, "Tienes una clase o reserva en ese horario.")
            return render(
                request,
                "siscad/profesor/reservar_aula.html",
                {
                    "aulas": aulas,
                    "horas": [],
                    "reservas": Reserva.objects.filter(profesor=profesor),
                    "fecha_seleccionada": None,
                    "dia_letra": None,
                },
            )

        # Crear la reserva
        reserva = Reserva.objects.create(
            profesor=profesor,
            aula=aula,
            fecha=fecha_seleccionada,
            curso=Curso.objects.first(),  # Puedes ajustarlo
        )

        # Asignar la hora
        hora_obj.tipo = "R"
        hora_obj.reserva = reserva
        hora_obj.save()

        messages.success(
            request,
            f" Reserva realizada con éxito para el aula {aula.nombre} el {fecha_seleccionada}.",
        )
        return redirect("reservar_aula")

    # -------------------- GET (mostrar horas disponibles) --------------------
    if request.GET.get("aula_id") and request.GET.get("fecha"):
        aula_id = request.GET.get("aula_id")
        fecha_str = request.GET.get("fecha")

        try:
            fecha_seleccionada = datetime.strptime(fecha_str, "%Y-%m-%d").date()
        except ValueError:
            messages.error(request, "Fecha inválida.")
            return redirect("reservar_aula")

        dia_semana = fecha_seleccionada.weekday()
        if dia_semana < 5:
            letras_dia = ["L", "M", "X", "J", "V"]
            dia_letra = letras_dia[dia_semana]
            aula = get_object_or_404(Aula, id=aula_id)
            horas_disponibles = Hora.objects.filter(
                aula=aula, dia=dia_letra, tipo__isnull=True
            ).order_by("hora_inicio")

    # -------------------- Contexto --------------------
    reservas_profesor = Reserva.objects.filter(profesor=profesor).select_related("aula")

    context = {
        "aulas": aulas,
        "horas": horas_disponibles,
        "reservas": reservas_profesor,
        "fecha_seleccionada": fecha_seleccionada,
        "dia_letra": dia_letra,
    }

    return render(request, "siscad/profesor/reservar_aula.html", context)


def cancelar_reserva(request, reserva_id):
    profesor = get_object_or_404(Profesor, email=request.session["email"])
    reserva = get_object_or_404(Reserva, id=reserva_id, profesor=profesor)

    ahora = timezone.localtime().time()
    hoy = timezone.localdate()

    if reserva.fecha > hoy or (
        reserva.fecha == hoy and all(h.hora_inicio > ahora for h in reserva.horas.all())
    ):
        for h in reserva.horas.all():
            h.tipo = None
            h.reserva = None
            h.save()

        reserva.delete()
        profesor.cantidad_reservas += 1
        profesor.save()

        messages.success(request, " Reserva cancelada exitosamente.")
    else:
        messages.error(
            request,
            " No puedes cancelar una reserva que ya ha comenzado o cuya fecha ha pasado.",
        )

    return redirect("reservar_aula")


def ver_cancelar_reservas(request):
    profesor = get_object_or_404(Profesor, email=request.session["email"])
    ahora = timezone.localtime()
    fecha_actual = ahora.date()
    hora_actual = ahora.time()

    # Eliminar reservas cuya fecha ya pasó
    reservas_pasadas = Reserva.objects.filter(fecha__lt=fecha_actual)
    for reserva in reservas_pasadas:
        reserva.delete()

    # Eliminar reservas del día actual que ya terminaron
    reservas_hoy = Reserva.objects.filter(fecha=fecha_actual)
    for reserva in reservas_hoy:
        horas_reserva = reserva.horas.all()
        if all(h.hora_fin < hora_actual for h in horas_reserva):
            reserva.delete()

    # Mostrar solo reservas activas
    reservas = Reserva.objects.filter(profesor=profesor).order_by("-fecha")

    context = {
        "reservas": reservas,
        "profesor": profesor,
    }
    return render(request, "siscad/profesor/cancelar_reservas.html", context)


def visualizar_asistencias_profesor(request):
    # 1. Obtener el profesor logueado
    if "email" not in request.session:
        return redirect("login")

    email = request.session["email"]

    try:
        profesor = Profesor.objects.get(email=email)
    except Profesor.DoesNotExist:
        return redirect("login")
    except Profesor.MultipleObjectsReturned:
        profesor = Profesor.objects.filter(email=email).first()

    # 2. Obtener cursos que enseña el profesor (teoría, práctica, laboratorio)
    cursos_ensenados = obtener_cursos_profesor(profesor)

    # 3. Procesar filtro por curso si se envió
    curso_seleccionado_id = request.GET.get("curso_id")
    curso_seleccionado = None
    asistencias_curso = []
    estadisticas_curso = {}

    if curso_seleccionado_id:
        try:
            curso_seleccionado = Curso.objects.get(id=curso_seleccionado_id)
            asistencias_curso = obtener_asistencias_curso_profesor(
                profesor, curso_seleccionado
            )
            estadisticas_curso = calcular_estadisticas_asistencias_profesor(
                asistencias_curso
            )
        except Curso.DoesNotExist:
            pass

    # 4. Preparar contexto
    context = {
        "profesor": profesor,
        "cursos_ensenados": cursos_ensenados,
        "curso_seleccionado": curso_seleccionado,
        "asistencias_curso": asistencias_curso,
        "estadisticas_curso": estadisticas_curso,
        "hoy": datetime.now().date(),
    }

    return render(
        request, "siscad/profesor/visualizar_asistencias_profesor.html", context
    )


def obtener_cursos_profesor(profesor):
    """
    Obtiene todos los cursos que enseña el profesor (teoría, práctica, laboratorio)
    """
    cursos = []

    # Cursos de teoría
    grupos_teoria = GrupoTeoria.objects.filter(profesor=profesor).select_related(
        "curso"
    )
    for grupo in grupos_teoria:
        cursos.append(
            {
                "id": grupo.curso.id,
                "nombre": grupo.curso.nombre,
                "tipo": "Teoría",
                "turno": grupo.turno,
                "grupo_id": grupo.id,
            }
        )

    # Cursos de práctica
    grupos_practica = GrupoPractica.objects.filter(profesor=profesor).select_related(
        "grupo_teoria__curso"
    )
    for grupo in grupos_practica:
        curso = grupo.grupo_teoria.curso
        # Solo agregar si no está ya en la lista
        if not any(c["id"] == curso.id and c["tipo"] == "Práctica" for c in cursos):
            cursos.append(
                {
                    "id": curso.id,
                    "nombre": curso.nombre,
                    "tipo": "Práctica",
                    "turno": grupo.turno,
                    "grupo_id": grupo.id,
                }
            )

    # Cursos de laboratorio
    grupos_lab = GrupoLaboratorio.objects.filter(profesor=profesor).select_related(
        "grupo_teoria__curso"
    )
    for grupo in grupos_lab:
        curso = grupo.grupo_teoria.curso
        # Solo agregar si no está ya en la lista
        if not any(c["id"] == curso.id and c["tipo"] == "Laboratorio" for c in cursos):
            cursos.append(
                {
                    "id": curso.id,
                    "nombre": curso.nombre,
                    "tipo": "Laboratorio",
                    "grupo_lab": grupo.grupo,
                    "grupo_id": grupo.id,
                }
            )

    return cursos


def obtener_asistencias_curso_profesor(profesor, curso):
    """
    Obtiene todas las asistencias del profesor para un curso específico
    """
    # Obtener asistencias relacionadas con este curso
    asistencias = (
        AsistenciaProfesor.objects.filter(profesor=profesor)
        .filter(
            Q(hora__grupo_teoria__curso=curso)
            | Q(hora__grupo_practica__grupo_teoria__curso=curso)
            | Q(hora__grupo_laboratorio__grupo_teoria__curso=curso)
        )
        .select_related(
            "hora__grupo_teoria__curso",
            "hora__grupo_practica__grupo_teoria__curso",
            "hora__grupo_laboratorio__grupo_teoria__curso",
        )
        .order_by("-fecha", "hora__hora_inicio")
    )

    # Formatear los datos para el template
    asistencias_formateadas = []
    for asistencia in asistencias:
        # Determinar tipo de clase y información
        tipo_clase = "Desconocido"
        detalle_clase = ""

        if asistencia.hora.grupo_teoria:
            tipo_clase = "Teoría"
            detalle_clase = f"T-{asistencia.hora.grupo_teoria.turno}"
        elif asistencia.hora.grupo_practica:
            tipo_clase = "Práctica"
            detalle_clase = f"P-{asistencia.hora.grupo_practica.turno}"
        elif asistencia.hora.grupo_laboratorio:
            tipo_clase = "Laboratorio"
            detalle_clase = f"L-{asistencia.hora.grupo_laboratorio.grupo}"

        asistencias_formateadas.append(
            {
                "fecha": asistencia.fecha,
                "dia_semana": asistencia.fecha.strftime("%A"),
                "hora_inicio": asistencia.hora.hora_inicio.strftime("%H:%M"),
                "hora_fin": asistencia.hora.hora_fin.strftime("%H:%M"),
                "tipo_clase": tipo_clase,
                "detalle_clase": detalle_clase,
                "estado": asistencia.estado,
                "estado_display": asistencia.get_estado_display(),
                "aula": asistencia.hora.aula.nombre
                if asistencia.hora.aula
                else "Sin aula",
                "es_pasado": asistencia.fecha < datetime.now().date(),
            }
        )

    return asistencias_formateadas


def calcular_estadisticas_asistencias_profesor(asistencias):
    """
    Calcula estadísticas de asistencias para un curso del profesor
    """
    if not asistencias:
        return {}

    total = len(asistencias)
    presentes = len([a for a in asistencias if a["estado"] == "P"])
    faltas = len([a for a in asistencias if a["estado"] == "F"])

    # Separar por tipo de clase
    teorias = [a for a in asistencias if a["tipo_clase"] == "Teoría"]
    practicas = [a for a in asistencias if a["tipo_clase"] == "Práctica"]
    laboratorios = [a for a in asistencias if a["tipo_clase"] == "Laboratorio"]

    # Calcular porcentajes
    porcentaje_asistencia = (presentes / total * 100) if total > 0 else 0

    return {
        "total": total,
        "presentes": presentes,
        "faltas": faltas,
        "porcentaje_asistencia": round(porcentaje_asistencia, 1),
        "teorias_total": len(teorias),
        "teorias_presentes": len([t for t in teorias if t["estado"] == "P"]),
        "practicas_total": len(practicas),
        "practicas_presentes": len([p for p in practicas if p["estado"] == "P"]),
        "laboratorios_total": len(laboratorios),
        "laboratorios_presentes": len([l for l in laboratorios if l["estado"] == "P"]),
    }


def registrar_asistencia(request):
    if "email" not in request.session:
        return redirect("login")

    profesor = get_object_or_404(Profesor, email=request.session["email"])

    # Reunir todos los grupos donde enseña
    grupos = []
    for gt in GrupoTeoria.objects.filter(profesor=profesor):
        grupos.append(
            {
                "id": gt.id,
                "nombre": f"{gt.curso.nombre} - Teoría {gt.turno}",
                "tipo": "teoria",
                "objeto": gt,
            }
        )
    for gp in GrupoPractica.objects.filter(profesor=profesor):
        grupos.append(
            {
                "id": gp.id,
                "nombre": f"{gp.grupo_teoria.curso.nombre} - Práctica {gp.turno}",
                "tipo": "practica",
                "objeto": gp,
            }
        )
    for gl in GrupoLaboratorio.objects.filter(profesor=profesor):
        grupos.append(
            {
                "id": gl.id,
                "nombre": f"{gl.grupo_teoria.curso.nombre} - Laboratorio {gl.grupo}",
                "tipo": "laboratorio",
                "objeto": gl,
            }
        )

    # Valores iniciales
    alumnos_asistencia = []
    fecha_seleccionada = timezone.localdate()
    hora_seleccionada = None
    grupo_seleccionado = None

    if request.method == "POST":
        grupo_id = request.POST.get("grupo_id")
        grupo_tipo = request.POST.get("grupo_tipo")
        fecha_input = request.POST.get("fecha")
        usar_hora_actual = request.POST.get("usar_hora_actual")

        # Procesar fecha
        if fecha_input:
            try:
                fecha_seleccionada = datetime.strptime(fecha_input, "%Y-%m-%d").date()
            except ValueError:
                fecha_seleccionada = timezone.localdate()
        else:
            fecha_seleccionada = timezone.localdate()

        # Obtener el día de la semana correctamente considerando timezone
        fecha_tz = timezone.make_aware(
            datetime.combine(fecha_seleccionada, datetime.min.time())
        )
        dia_numero = (
            fecha_tz.weekday()
        )  # 0=Lunes, 1=Martes, 2=Miércoles, 3=Jueves, 4=Viernes
        dias_map = {
            0: "L",  # Lunes
            1: "M",  # Martes
            2: "X",  # Miércoles
            3: "J",  # Jueves
            4: "V",  # Viernes
        }
        dia_codigo = dias_map.get(dia_numero, "")

        print(
            f"DEBUG: Fecha: {fecha_seleccionada}, Día número: {dia_numero}, Día código: {dia_codigo}"
        )

        # Obtener grupo seleccionado
        if grupo_tipo and grupo_id:
            if grupo_tipo == "teoria":
                grupo_seleccionado = get_object_or_404(GrupoTeoria, id=grupo_id)
            elif grupo_tipo == "practica":
                grupo_seleccionado = get_object_or_404(GrupoPractica, id=grupo_id)
            elif grupo_tipo == "laboratorio":
                grupo_seleccionado = get_object_or_404(GrupoLaboratorio, id=grupo_id)

        # Determinar hora seleccionada automáticamente basada en el grupo y día
        if grupo_seleccionado and dia_codigo:
            hora_seleccionada = obtener_hora_para_grupo(
                grupo_seleccionado, grupo_tipo, dia_codigo
            )
            print(f"DEBUG: Hora encontrada para el grupo: {hora_seleccionada}")

        # Si se usa hora actual, buscar la hora correspondiente
        if usar_hora_actual and grupo_seleccionado and dia_codigo:
            ahora = timezone.localtime()
            hora_actual = ahora.time()
            fecha_seleccionada = ahora.date()

            # Recalcular día de la semana con la fecha actual
            dia_numero_actual = ahora.weekday()
            dia_codigo_actual = dias_map.get(dia_numero_actual, "")

            if dia_codigo_actual:
                # Buscar hora actual para el grupo
                hora_seleccionada = obtener_hora_actual_para_grupo(
                    grupo_seleccionado, grupo_tipo, dia_codigo_actual, hora_actual
                )
                print(f"DEBUG: Hora actual encontrada: {hora_seleccionada}")

        # Obtener alumnos según tipo de grupo
        if grupo_tipo and grupo_id and hora_seleccionada:
            alumnos = obtener_alumnos_para_grupo(grupo_seleccionado, grupo_tipo)
            print(f"DEBUG: Alumnos encontrados: {alumnos.count()}")

            # Cargar asistencias existentes
            asistencias_existentes = AsistenciaAlumno.objects.filter(
                fecha=fecha_seleccionada, hora=hora_seleccionada, alumno__in=alumnos
            )
            asistencia_dict = {a.alumno_id: a.estado for a in asistencias_existentes}

            # Crear lista de alumnos con su estado de asistencia
            alumnos_asistencia = []
            for alumno in alumnos:
                estado = asistencia_dict.get(alumno.id, "F")  # Por defecto Falta
                alumnos_asistencia.append(
                    {"alumno": alumno, "estado": estado, "id": alumno.id}
                )

            print(f"DEBUG: Alumnos para asistencia: {len(alumnos_asistencia)}")

        # Guardar asistencias
        if (
            "guardar_asistencia" in request.POST
            and grupo_seleccionado
            and hora_seleccionada
        ):
            for alumno_data in alumnos_asistencia:
                alumno_id = alumno_data["alumno"].id
                estado = request.POST.get(f"asistencia_{alumno_id}", "F")

                AsistenciaAlumno.objects.update_or_create(
                    alumno_id=alumno_id,
                    fecha=fecha_seleccionada,
                    hora=hora_seleccionada,
                    defaults={"estado": estado},
                )

            print("DEBUG: Asistencias guardadas correctamente")
            # Redirigir para evitar reenvío del formulario
            return redirect("registrar_asistencia")

    context = {
        "grupos": grupos,
        "alumnos_asistencia": alumnos_asistencia,
        "fecha": fecha_seleccionada,
        "hora": hora_seleccionada,
        "grupo_seleccionado": grupo_seleccionado,
    }
    return render(request, "siscad/profesor/registrar_asistencia.html", context)


def obtener_hora_para_grupo(grupo, grupo_tipo, dia_codigo):
    """
    Obtiene la hora automáticamente para un grupo en un día específico
    """
    try:
        if grupo_tipo == "teoria":
            # Buscar horas de teoría para este grupo en el día específico
            horas = Hora.objects.filter(
                grupo_teoria=grupo,
                dia=dia_codigo,
                tipo="T",  # Teoría
            ).order_by("hora_inicio")

        elif grupo_tipo == "practica":
            # Buscar horas de práctica para este grupo en el día específico
            horas = Hora.objects.filter(
                grupo_practica=grupo,
                dia=dia_codigo,
                tipo="P",  # Práctica
            ).order_by("hora_inicio")

        elif grupo_tipo == "laboratorio":
            # Buscar horas de laboratorio para este grupo en el día específico
            horas = Hora.objects.filter(
                grupo_laboratorio=grupo,
                dia=dia_codigo,
                tipo="L",  # Laboratorio
            ).order_by("hora_inicio")
        else:
            return None

        # Si hay múltiples horas, tomar la primera (podrías implementar lógica más compleja aquí)
        if horas.exists():
            return horas.first()

        # Si no encuentra hora específica, buscar cualquier hora para ese grupo y día
        if grupo_tipo == "teoria":
            horas = Hora.objects.filter(grupo_teoria=grupo, dia=dia_codigo)
        elif grupo_tipo == "practica":
            horas = Hora.objects.filter(grupo_practica=grupo, dia=dia_codigo)
        elif grupo_tipo == "laboratorio":
            horas = Hora.objects.filter(grupo_laboratorio=grupo, dia=dia_codigo)

        return horas.first() if horas.exists() else None

    except Exception as e:
        print(f"ERROR obteniendo hora para grupo: {e}")
        return None


def obtener_hora_actual_para_grupo(grupo, grupo_tipo, dia_codigo, hora_actual):
    """
    Obtiene la hora actual para un grupo basado en la hora del sistema
    """
    try:
        if grupo_tipo == "teoria":
            horas = Hora.objects.filter(
                grupo_teoria=grupo,
                dia=dia_codigo,
                hora_inicio__lte=hora_actual,
                hora_fin__gte=hora_actual,
            )
        elif grupo_tipo == "practica":
            horas = Hora.objects.filter(
                grupo_practica=grupo,
                dia=dia_codigo,
                hora_inicio__lte=hora_actual,
                hora_fin__gte=hora_actual,
            )
        elif grupo_tipo == "laboratorio":
            horas = Hora.objects.filter(
                grupo_laboratorio=grupo,
                dia=dia_codigo,
                hora_inicio__lte=hora_actual,
                hora_fin__gte=hora_actual,
            )
        else:
            return None

        # Si encuentra una hora que coincide con la hora actual, usarla
        if horas.exists():
            return horas.first()

        # Si no encuentra hora actual, buscar la primera hora del día para ese grupo
        return obtener_hora_para_grupo(grupo, grupo_tipo, dia_codigo)

    except Exception as e:
        print(f"ERROR obteniendo hora actual para grupo: {e}")
        return None


def obtener_alumnos_para_grupo(grupo, grupo_tipo):
    """
    Obtiene los alumnos para un grupo específico
    """
    try:
        if grupo_tipo == "teoria":
            # Alumnos matriculados en el curso con el mismo turno
            alumnos = Alumno.objects.filter(
                matriculas_curso__curso=grupo.curso, matriculas_curso__turno=grupo.turno
            ).distinct()

        elif grupo_tipo == "practica":
            # Alumnos matriculados en el curso con el mismo turno
            alumnos = Alumno.objects.filter(
                matriculas_curso__curso=grupo.grupo_teoria.curso,
                matriculas_curso__turno=grupo.turno,
            ).distinct()

        elif grupo_tipo == "laboratorio":
            # Alumnos matriculados en este laboratorio
            alumnos = Alumno.objects.filter(
                matriculas_laboratorio__grupo_laboratorio=grupo
            ).distinct()
        else:
            alumnos = Alumno.objects.none()

        return alumnos

    except Exception as e:
        print(f"ERROR obteniendo alumnos para grupo: {e}")
        return Alumno.objects.none()


def ingresar_notas(request):
    if "email" not in request.session:
        return redirect("login")

    profesor = get_object_or_404(Profesor, email=request.session["email"])
    grupos_teoria = GrupoTeoria.objects.filter(profesor=profesor)

    grupo_seleccionado = None
    alumnos_grupo = []
    notas_data = []

    if request.method == "POST":
        grupo_id = request.POST.get("grupo_id")

        if grupo_id:
            grupo_seleccionado = get_object_or_404(
                GrupoTeoria, id=grupo_id, profesor=profesor
            )

            alumnos_grupo = MatriculaCurso.objects.filter(
                curso=grupo_seleccionado.curso, turno=grupo_seleccionado.turno
            ).select_related("alumno")

            if "guardar_manual" in request.POST:
                procesar_notas_manual(request, grupo_seleccionado, alumnos_grupo)
                messages.success(request, "Notas guardadas correctamente")
                return redirect("ingresar_notas")

            elif "procesar_excel" in request.POST and request.FILES.get(
                "archivo_excel"
            ):
                archivo = request.FILES["archivo_excel"]
                resultado = procesar_excel_notas(archivo, grupo_seleccionado)
                if resultado["success"]:
                    messages.success(request, resultado["message"])
                    return redirect("ingresar_notas")
                else:
                    messages.error(request, resultado["message"])

            notas_data = preparar_datos_notas(alumnos_grupo, grupo_seleccionado)

    context = {
        "grupos_teoria": grupos_teoria,
        "grupo_seleccionado": grupo_seleccionado,
        "alumnos_grupo": alumnos_grupo,
        "notas_data": notas_data,
    }

    return render(request, "siscad/profesor/ingresar_notas.html", context)


def procesar_notas_manual(request, grupo_teoria, alumnos_grupo):
    """
    Procesa las notas ingresadas manualmente en el formulario
    """
    curso = grupo_teoria.curso

    with transaction.atomic():
        for alumno_matricula in alumnos_grupo:
            alumno = alumno_matricula.alumno

            # Procesar notas continuas (solo si el curso las tiene configuradas)
            if any(
                [
                    curso.peso_continua_1 > 0,
                    curso.peso_continua_2 > 0,
                    curso.peso_continua_3 > 0,
                ]
            ):
                for periodo in [1, 2, 3]:
                    peso_continua = getattr(curso, f"peso_continua_{periodo}", 0)
                    # Solo procesar si el periodo tiene peso configurado
                    if peso_continua > 0:
                        campo_continua = f"continua_{alumno.id}_{periodo}"
                        valor_continua = request.POST.get(campo_continua)

                        if valor_continua and valor_continua.strip():
                            try:
                                valor_float = float(valor_continua)
                                nota, created = Nota.objects.get_or_create(
                                    alumno=alumno,
                                    curso=curso,
                                    tipo="C",
                                    periodo=periodo,
                                    defaults={
                                        "valor": valor_float,
                                        "peso": peso_continua,
                                    },
                                )
                                if not created:
                                    nota.valor = valor_float
                                    nota.peso = peso_continua
                                    nota.save()
                            except ValueError:
                                continue
                        # Si el campo está vacío, establecer como -1
                        elif valor_continua == "":
                            Nota.objects.filter(
                                alumno=alumno, curso=curso, tipo="C", periodo=periodo
                            ).update(valor=-1)

            # Procesar 3 parciales (solo si el curso los tiene configurados)
            if any(
                [
                    curso.peso_parcial_1 > 0,
                    curso.peso_parcial_2 > 0,
                    curso.peso_parcial_3 > 0,
                ]
            ):
                for periodo_parcial in [1, 2, 3]:
                    peso_parcial = getattr(curso, f"peso_parcial_{periodo_parcial}", 0)
                    # Solo procesar si el periodo tiene peso configurado
                    if peso_parcial > 0:
                        campo_parcial = f"parcial_{alumno.id}_{periodo_parcial}"
                        valor_parcial = request.POST.get(campo_parcial)

                        if valor_parcial and valor_parcial.strip():
                            try:
                                valor_float = float(valor_parcial)
                                nota, created = Nota.objects.get_or_create(
                                    alumno=alumno,
                                    curso=curso,
                                    tipo="P",
                                    periodo=periodo_parcial,
                                    defaults={
                                        "valor": valor_float,
                                        "peso": peso_parcial,
                                    },
                                )
                                if not created:
                                    nota.valor = valor_float
                                    nota.peso = peso_parcial
                                    nota.save()
                            except ValueError:
                                continue
                        # Si el campo está vacío, establecer como -1
                        elif valor_parcial == "":
                            Nota.objects.filter(
                                alumno=alumno,
                                curso=curso,
                                tipo="P",
                                periodo=periodo_parcial,
                            ).update(valor=-1)

            # Procesar sustitutorio (solo si el curso tiene al menos 2 parciales)
            if curso.peso_parcial_1 > 0 and curso.peso_parcial_2 > 0:
                campo_sustitutorio = f"sustitutorio_{alumno.id}"
                valor_sustitutorio = request.POST.get(campo_sustitutorio)

                if valor_sustitutorio and valor_sustitutorio.strip():
                    try:
                        valor_float = float(valor_sustitutorio)
                        nota, created = Nota.objects.get_or_create(
                            alumno=alumno,
                            curso=curso,
                            tipo="S",
                            periodo=1,
                            defaults={"valor": valor_float, "peso": 1},
                        )
                        if not created:
                            nota.valor = valor_float
                            nota.save()
                    except ValueError:
                        continue
                # Si el campo está vacío, establecer como -1
                elif valor_sustitutorio == "":
                    Nota.objects.filter(
                        alumno=alumno, curso=curso, tipo="S", periodo=1
                    ).update(valor=-1)


def procesar_excel_notas(archivo, grupo_teoria):
    """
    Procesa un archivo Excel para cargar notas masivamente - VERSIÓN CON -1
    """
    try:
        # Leer el archivo más rápido sin procesamiento extra
        if archivo.name.endswith(".xlsx"):
            df = pd.read_excel(archivo, dtype={"dni_alumno": str, "valor": object})
        elif archivo.name.endswith(".csv"):
            df = pd.read_csv(archivo, dtype={"dni_alumno": str, "valor": object})
        else:
            return {"success": False, "message": "Formato de archivo no soportado"}

        # Validar columnas rápidamente
        columnas_requeridas = ["dni_alumno", "tipo_nota", "periodo", "valor"]
        if not all(col in df.columns for col in columnas_requeridas):
            return {"success": False, "message": "Faltan columnas requeridas"}

        # Limpiar y preparar datos
        df = df.copy()
        df["dni_alumno"] = df["dni_alumno"].astype(str).str.strip()
        df["tipo_nota"] = df["tipo_nota"].astype(str).str.strip().str.upper()
        df["periodo"] = (
            pd.to_numeric(df["periodo"], errors="coerce").fillna(0).astype(int)
        )

        # Manejar valores vacíos: convertir a -1
        df["valor"] = pd.to_numeric(df["valor"], errors="coerce")
        df["valor"] = df["valor"].fillna(-1)  # Valores vacíos se convierten a -1

        # Filtrar valores válidos (>= 0 y <= 20) y -1 para "no asignado"
        df = df[((df["valor"] >= 0) & (df["valor"] <= 20)) | (df["valor"] == -1)]

        # Filtrar tipos de nota válidos y periodos válidos
        df = df[df["tipo_nota"].isin(["C", "P", "S"]) & df["periodo"].isin([1, 2, 3])]

        curso = grupo_teoria.curso

        # Pre-cache de datos para mejor performance
        alumnos_dni = set(df["dni_alumno"].unique())
        alumnos_dict = {
            alumno.dni: alumno for alumno in Alumno.objects.filter(dni__in=alumnos_dni)
        }

        # Verificar matriculas en una sola consulta
        matriculas_existentes = set(
            MatriculaCurso.objects.filter(
                alumno__dni__in=alumnos_dni, curso=curso, turno=grupo_teoria.turno
            ).values_list("alumno__dni", flat=True)
        )

        # Filtrar solo alumnos matriculados
        df = df[df["dni_alumno"].isin(matriculas_existentes)]

        if df.empty:
            return {
                "success": False,
                "message": "No hay datos válidos para procesar",
            }

        # Preparar datos para bulk_create y bulk_update
        notas_a_procesar = []

        # Agrupar por alumno y tipo para procesamiento más eficiente
        for (dni_alumno, tipo_nota, periodo), group_df in df.groupby(
            ["dni_alumno", "tipo_nota", "periodo"]
        ):
            if not group_df.empty:
                valor = group_df["valor"].iloc[0]  # Tomar el primer valor

                # Validar configuración del curso
                if tipo_nota == "C":
                    peso = getattr(curso, f"peso_continua_{periodo}", 0)
                    if peso == 0:
                        continue
                elif tipo_nota == "P":
                    peso = getattr(curso, f"peso_parcial_{periodo}", 0)
                    if peso == 0:
                        continue
                else:  # Sustitutorio
                    if not (curso.peso_parcial_1 > 0 and curso.peso_parcial_2 > 0):
                        continue
                    peso = 1

                alumno = alumnos_dict.get(dni_alumno)
                if not alumno:
                    continue

                notas_a_procesar.append(
                    {
                        "alumno": alumno,
                        "tipo": tipo_nota,
                        "periodo": periodo,
                        "valor": valor,
                        "peso": peso,
                    }
                )

        if not notas_a_procesar:
            return {"success": False, "message": "No hay notas válidas para procesar"}

        # Procesamiento masivo con bulk operations
        with transaction.atomic():
            # Obtener notas existentes para evitar duplicados
            notas_existentes = Nota.objects.filter(
                alumno__in=[n["alumno"] for n in notas_a_procesar], curso=curso
            ).select_related("alumno")

            # Crear diccionario de notas existentes para búsqueda rápida
            existentes_dict = {}
            for nota in notas_existentes:
                key = (nota.alumno.dni, nota.tipo, nota.periodo)
                existentes_dict[key] = nota

            # Separar en crear y actualizar
            notas_para_crear = []
            notas_para_actualizar = []

            for nota_data in notas_a_procesar:
                key = (nota_data["alumno"].dni, nota_data["tipo"], nota_data["periodo"])
                if key in existentes_dict:
                    nota_existente = existentes_dict[key]
                    nota_existente.valor = nota_data["valor"]
                    nota_existente.peso = nota_data["peso"]
                    notas_para_actualizar.append(nota_existente)
                else:
                    notas_para_crear.append(
                        Nota(
                            alumno=nota_data["alumno"],
                            curso=curso,
                            tipo=nota_data["tipo"],
                            periodo=nota_data["periodo"],
                            valor=nota_data["valor"],
                            peso=nota_data["peso"],
                        )
                    )

            # Ejecutar operaciones masivas
            if notas_para_crear:
                Nota.objects.bulk_create(notas_para_crear, batch_size=1000)

            if notas_para_actualizar:
                Nota.objects.bulk_update(
                    notas_para_actualizar, ["valor", "peso"], batch_size=1000
                )

            total_procesadas = len(notas_para_crear) + len(notas_para_actualizar)

        return {
            "success": True,
            "message": f"Se procesaron {total_procesadas} registros correctamente "
            f"({len(notas_para_crear)} nuevos, {len(notas_para_actualizar)} actualizados)",
        }

    except Exception as e:
        return {"success": False, "message": f"Error al procesar el archivo: {str(e)}"}


def preparar_datos_notas(alumnos_grupo, grupo_teoria):
    """
    Prepara los datos de notas para mostrar en la tabla
    """
    datos = []
    curso = grupo_teoria.curso

    for matricula in alumnos_grupo:
        alumno = matricula.alumno

        # Obtener todas las notas del alumno en este curso
        notas_alumno = Nota.objects.filter(alumno=alumno, curso=curso)

        # Organizar notas (-1 se muestra como vacío)
        notas_continua = {1: None, 2: None, 3: None}
        notas_parcial = {1: None, 2: None, 3: None}
        nota_sustitutorio = None

        for nota in notas_alumno:
            if nota.tipo == "C" and nota.periodo in [1, 2, 3]:
                # Mostrar solo valores >= 0, -1 se muestra como None/vacío
                notas_continua[nota.periodo] = nota.valor if nota.valor >= 0 else None
            elif nota.tipo == "P" and nota.periodo in [1, 2, 3]:
                notas_parcial[nota.periodo] = nota.valor if nota.valor >= 0 else None
            elif nota.tipo == "S":
                nota_sustitutorio = nota.valor if nota.valor >= 0 else None

        # Determinar qué tipos de notas mostrar según configuración del curso
        curso_tiene_continua = any(
            [
                curso.peso_continua_1 > 0,
                curso.peso_continua_2 > 0,
                curso.peso_continua_3 > 0,
            ]
        )

        curso_tiene_parciales = any(
            [
                curso.peso_parcial_1 > 0,
                curso.peso_parcial_2 > 0,
                curso.peso_parcial_3 > 0,
            ]
        )

        curso_tiene_sustitutorio = curso.peso_parcial_1 > 0 and curso.peso_parcial_2 > 0

        datos.append(
            {
                "alumno": alumno,
                "matricula": matricula,
                "notas_continua": notas_continua,
                "notas_parcial": notas_parcial,
                "nota_sustitutorio": nota_sustitutorio,
                "curso_tiene_continua": curso_tiene_continua,
                "curso_tiene_parciales": curso_tiene_parciales,
                "curso_tiene_sustitutorio": curso_tiene_sustitutorio,
                "pesos_continua": {
                    1: curso.peso_continua_1,
                    2: curso.peso_continua_2,
                    3: curso.peso_continua_3,
                },
                "pesos_parcial": {
                    1: curso.peso_parcial_1,
                    2: curso.peso_parcial_2,
                    3: curso.peso_parcial_3,
                },
            }
        )

    return datos


def descargar_plantilla_excel(request, grupo_id):
    """
    Vista para descargar una plantilla Excel para cargar notas CON DATOS EXISTENTES
    """
    grupo_teoria = get_object_or_404(GrupoTeoria, id=grupo_id)
    curso = grupo_teoria.curso

    # Crear DataFrame con la estructura requerida
    data = {
        "dni_alumno": [],
        "tipo_nota": [],  # C: Continua, P: Parcial, S: Sustitutorio
        "periodo": [],  # 1, 2, 3 para continua/parcial, 1 para sustitutorio
        "valor": [],
    }

    # Agregar alumnos del grupo
    alumnos_grupo = MatriculaCurso.objects.filter(
        curso=grupo_teoria.curso, turno=grupo_teoria.turno
    ).select_related("alumno")

    # Pre-cargar todas las notas de estos alumnos en este curso para mejor performance
    alumnos_ids = [matricula.alumno.id for matricula in alumnos_grupo]
    notas_existentes = Nota.objects.filter(
        alumno_id__in=alumnos_ids, curso=curso
    ).select_related("alumno")

    # Organizar notas por alumno para acceso rápido
    notas_por_alumno = {}
    for nota in notas_existentes:
        if nota.alumno_id not in notas_por_alumno:
            notas_por_alumno[nota.alumno_id] = []
        notas_por_alumno[nota.alumno_id].append(nota)

    for matricula in alumnos_grupo:
        alumno = matricula.alumno
        notas_alumno = notas_por_alumno.get(alumno.id, [])

        # Organizar notas del alumno por tipo y periodo
        notas_organizadas = {
            "C": {1: None, 2: None, 3: None},
            "P": {1: None, 2: None, 3: None},
            "S": {1: None},
        }

        for nota in notas_alumno:
            if nota.tipo in ["C", "P"] and nota.periodo in [1, 2, 3]:
                # Solo mostrar valores >= 0, -1 se muestra como vacío
                notas_organizadas[nota.tipo][nota.periodo] = (
                    nota.valor if nota.valor >= 0 else None
                )
            elif nota.tipo == "S" and nota.periodo == 1:
                notas_organizadas["S"][1] = nota.valor if nota.valor >= 0 else None

        # Agregar filas para notas continuas (solo si el curso las tiene)
        if any(
            [
                curso.peso_continua_1 > 0,
                curso.peso_continua_2 > 0,
                curso.peso_continua_3 > 0,
            ]
        ):
            for periodo in [1, 2, 3]:
                peso_continua = getattr(curso, f"peso_continua_{periodo}", 0)
                if peso_continua > 0:
                    data["dni_alumno"].append(alumno.dni)
                    data["tipo_nota"].append("C")
                    data["periodo"].append(periodo)
                    # Usar valor existente si existe y es >= 0, sino vacío
                    valor_existente = notas_organizadas["C"][periodo]
                    data["valor"].append(
                        valor_existente if valor_existente is not None else ""
                    )

        # Agregar filas para notas parciales (solo si el curso las tiene)
        if any(
            [
                curso.peso_parcial_1 > 0,
                curso.peso_parcial_2 > 0,
                curso.peso_parcial_3 > 0,
            ]
        ):
            for periodo in [1, 2, 3]:
                peso_parcial = getattr(curso, f"peso_parcial_{periodo}", 0)
                if peso_parcial > 0:
                    data["dni_alumno"].append(alumno.dni)
                    data["tipo_nota"].append("P")
                    data["periodo"].append(periodo)
                    # Usar valor existente si existe y es >= 0, sino vacío
                    valor_existente = notas_organizadas["P"][periodo]
                    data["valor"].append(
                        valor_existente if valor_existente is not None else ""
                    )

        # Agregar fila para sustitutorio (solo si el curso lo permite)
        if curso.peso_parcial_1 > 0 and curso.peso_parcial_2 > 0:
            data["dni_alumno"].append(alumno.dni)
            data["tipo_nota"].append("S")
            data["periodo"].append(1)
            # Usar valor existente si existe y es >= 0, sino vacío
            valor_existente = notas_organizadas["S"][1]
            data["valor"].append(valor_existente if valor_existente is not None else "")

    df = pd.DataFrame(data)

    # Crear un libro de Excel con formato mejorado
    output = io.BytesIO()

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        # Hoja principal con datos
        df.to_excel(writer, sheet_name="Plantilla_Notas", index=False)

        # Obtener la hoja para aplicar formato
        workbook = writer.book
        worksheet = writer.sheets["Plantilla_Notas"]

        # Aplicar formato a las columnas
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        # Estilo para encabezados
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_fill = PatternFill(
            start_color="366092", end_color="366092", fill_type="solid"
        )
        header_alignment = Alignment(horizontal="center", vertical="center")
        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # Aplicar estilo a los encabezados
        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border

        # Aplicar bordes a todas las celdas con datos
        for row in worksheet.iter_rows(
            min_row=2, max_row=worksheet.max_row, min_col=1, max_col=4
        ):
            for cell in row:
                cell.border = border

        # Ajustar ancho de columnas
        column_widths = {
            "A": 15,  # dni_alumno
            "B": 12,  # tipo_nota
            "C": 10,  # periodo
            "D": 12,  # valor
        }

        for col, width in column_widths.items():
            worksheet.column_dimensions[col].width = width

        # Aplicar formato de texto a las columnas
        for row in range(2, worksheet.max_row + 1):
            # DNI como texto
            worksheet[f"A{row}"].number_format = "@"
            # Valor con formato numérico (2 decimales)
            worksheet[f"D{row}"].number_format = "0.00"

        # Hoja de instrucciones
        instrucciones_data = {
            "Instrucciones": [
                "1. NO modifique la estructura de las columnas",
                "2. Los tipos de nota son: C=Continua, P=Parcial, S=Sustitutorio",
                "3. Los periodos válidos son: 1, 2, 3",
                "4. Las notas deben estar entre 0 y 20",
                "5. Deje vacío si no desea modificar una nota existente",
                "6. Para eliminar una nota, déjela vacía en el Excel",
                "7. Mantenga el formato del DNI sin espacios",
                "8. Los valores vacíos se interpretarán como 'no evaluado'",
                "9. Solo se procesarán las notas que coincidan con la configuración del curso",
                "10. El sustitutorio solo aplica si el curso tiene Parcial 1 y Parcial 2 configurados",
            ]
        }
        df_instrucciones = pd.DataFrame(instrucciones_data)
        df_instrucciones.to_excel(writer, sheet_name="Instrucciones", index=False)

        # Formato hoja de instrucciones
        worksheet_inst = writer.sheets["Instrucciones"]

        # Aplicar formato a la hoja de instrucciones
        for cell in worksheet_inst[1]:
            cell.font = Font(bold=True, color="FFFFFF", size=12)
            cell.fill = PatternFill(
                start_color="28a745", end_color="28a745", fill_type="solid"
            )
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border

        # Ajustar ancho de columna de instrucciones
        worksheet_inst.column_dimensions["A"].width = 80

        # Aplicar bordes y formato a las instrucciones
        for row in range(2, worksheet_inst.max_row + 1):
            worksheet_inst[f"A{row}"].border = border
            worksheet_inst[f"A{row}"].alignment = Alignment(
                wrap_text=True, vertical="center"
            )

        # Hoja de configuración del curso
        config_data = {
            "Configuración del Curso": [
                f"Curso: {curso.nombre}",
                f"Código: {curso.codigo}",
                f"Grupo: {grupo_teoria.get_turno_display()}",
                "",
                "Pesos de Evaluación:",
                f"Continua P1: {curso.peso_continua_1}%",
                f"Continua P2: {curso.peso_continua_2}%",
                f"Continua P3: {curso.peso_continua_3}%",
                f"Parcial 1: {curso.peso_parcial_1}%",
                f"Parcial 2: {curso.peso_parcial_2}%",
                f"Parcial 3: {curso.peso_parcial_3}%",
                "",
                "Sustitutorio: "
                + (
                    "HABILITADO (reemplaza la nota más baja entre P1 y P2)"
                    if curso.peso_parcial_1 > 0 and curso.peso_parcial_2 > 0
                    else "NO HABILITADO"
                ),
                "",
                f"Total alumnos en el grupo: {len(alumnos_grupo)}",
                f"Fecha de generación: {timezone.now().strftime('%Y-%m-%d %H:%M')}",
            ]
        }
        df_config = pd.DataFrame(config_data)
        df_config.to_excel(writer, sheet_name="Configuración", index=False)

        # Formato hoja de configuración
        worksheet_config = writer.sheets["Configuración"]
        worksheet_config.column_dimensions["A"].width = 60

        # Aplicar formato a la hoja de configuración
        for cell in worksheet_config[1]:
            cell.font = Font(bold=True, color="FFFFFF", size=12)
            cell.fill = PatternFill(
                start_color="6f42c1", end_color="6f42c1", fill_type="solid"
            )
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border

        # Aplicar bordes y formato a la configuración
        for row in range(2, worksheet_config.max_row + 1):
            worksheet_config[f"A{row}"].border = border
            worksheet_config[f"A{row}"].alignment = Alignment(
                wrap_text=True, vertical="center"
            )

        # Resaltar filas importantes en configuración
        important_rows = [5, 12]  # Filas de "Pesos de Evaluación" y "Sustitutorio"
        for row in important_rows:
            if row <= worksheet_config.max_row:
                worksheet_config[f"A{row}"].font = Font(bold=True, color="2c3e50")

    output.seek(0)

    # Crear respuesta HTTP
    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = (
        f'attachment; filename="plantilla_notas_{curso.codigo}_{grupo_teoria.get_turno_display()}_{timezone.now().strftime("%Y%m%d")}.xlsx"'
    )

    return response


def calcular_nota_final(alumno, curso):
    """
    Calcula la nota final considerando sustitutorio y pesos dinámicos
    El sustitutorio reemplaza la nota más baja de los parciales 1 y 2
    IGNORA las notas que son -1 (no asignadas)
    """
    # Obtener todas las notas
    notas = Nota.objects.filter(alumno=alumno, curso=curso)

    # Separar por tipo, excluyendo -1
    continuas = {
        n.periodo: n
        for n in notas
        if n.tipo == "C" and n.valor is not None and n.valor >= 0
    }
    parciales = {
        n.periodo: n
        for n in notas
        if n.tipo == "P" and n.valor is not None and n.valor >= 0
    }
    sustitutorio = next(
        (n for n in notas if n.tipo == "S" and n.valor is not None and n.valor >= 0),
        None,
    )

    # Aplicar sustitutorio si existe y hay al menos 2 parciales con nota válida
    if (
        sustitutorio
        and sustitutorio.valor is not None
        and sustitutorio.valor >= 0
        and 1 in parciales
        and 2 in parciales
    ):
        # Encontrar el parcial más bajo entre P1 y P2
        if parciales[1].valor <= parciales[2].valor:
            # Reemplazar P1 temporalmente para el cálculo
            nota_original_p1 = parciales[1].valor
            parciales[1].valor = sustitutorio.valor
        else:
            # Reemplazar P2 temporalmente para el cálculo
            nota_original_p2 = parciales[2].valor
            parciales[2].valor = sustitutorio.valor

    # Calcular promedio ponderado
    total_puntos = 0
    total_pesos = 0

    # Sumar continuas (solo si tienen peso configurado y valor >= 0)
    for periodo in [1, 2, 3]:
        if periodo in continuas:
            peso = getattr(curso, f"peso_continua_{periodo}", 0)
            if peso > 0 and continuas[periodo].valor >= 0:
                total_puntos += continuas[periodo].valor * peso
                total_pesos += peso

    # Sumar parciales (solo si tienen peso configurado y valor >= 0)
    for periodo in [1, 2, 3]:
        if periodo in parciales:
            peso = getattr(curso, f"peso_parcial_{periodo}", 0)
            if peso > 0 and parciales[periodo].valor >= 0:
                total_puntos += parciales[periodo].valor * peso
                total_pesos += peso

    # Solo calcular si hay al menos una nota válida
    if total_pesos == 0:
        return -1  # Retorna -1 si no hay notas válidas

    nota_final = total_puntos / total_pesos

    # Restaurar notas originales si se aplicó sustitutorio
    if (
        sustitutorio
        and sustitutorio.valor is not None
        and sustitutorio.valor >= 0
        and 1 in parciales
        and 2 in parciales
    ):
        if "nota_original_p1" in locals():
            parciales[1].valor = nota_original_p1
        elif "nota_original_p2" in locals():
            parciales[2].valor = nota_original_p2

    return round(nota_final, 2)


def revisar_estadisticas(request):
    if "email" not in request.session:
        return redirect("login")

    profesor = get_object_or_404(Profesor, email=request.session["email"])

    grupos_teoria = GrupoTeoria.objects.filter(profesor=profesor)

    grupo_seleccionado = None
    estadisticas_generales = None
    estadisticas_detalladas = None
    estadisticas_avance = None
    alumnos_grupo = []

    if request.method == "POST":
        grupo_id = request.POST.get("grupo_id")

        if grupo_id:
            grupo_seleccionado = get_object_or_404(
                GrupoTeoria, id=grupo_id, profesor=profesor
            )

            # Obtener alumnos del grupo
            alumnos_grupo = MatriculaCurso.objects.filter(
                curso=grupo_seleccionado.curso, turno=grupo_seleccionado.turno
            ).select_related("alumno")

            # Procesar subida de exámenes
            if "subir_examenes" in request.POST and request.FILES.getlist(
                "archivos_examen"
            ):
                procesar_examenes(request, grupo_seleccionado)
                messages.success(request, "Exámenes subidos correctamente")
                return redirect("revisar_estadisticas")

            # Descargar Excel
            elif "descargar_excel" in request.POST:
                return descargar_estadisticas_excel(
                    request, grupo_seleccionado, alumnos_grupo
                )

            # Generar estadísticas
            else:
                estadisticas_generales = calcular_estadisticas_generales(
                    grupo_seleccionado, alumnos_grupo
                )
                estadisticas_detalladas = calcular_estadisticas_detalladas(
                    grupo_seleccionado, alumnos_grupo
                )
                estadisticas_avance = calcular_estadisticas_avance(
                    grupo_seleccionado, alumnos_grupo
                )

    context = {
        "grupos_teoria": grupos_teoria,
        "grupo_seleccionado": grupo_seleccionado,
        "estadisticas_generales": estadisticas_generales,
        "estadisticas_detalladas": estadisticas_detalladas,
        "estadisticas_avance": estadisticas_avance,
        "alumnos_grupo": alumnos_grupo,
    }

    return render(request, "siscad/profesor/revisar_estadisticas.html", context)


def calcular_estadisticas_generales(grupo_teoria, alumnos_grupo):
    """
    Calcula estadísticas generales del grupo - VERSIÓN CON -1
    """
    curso = grupo_teoria.curso
    notas_finales = []

    for matricula in alumnos_grupo:
        alumno = matricula.alumno
        nota_final = calcular_nota_final(alumno, curso)
        # Solo considerar notas válidas (>= 0)
        if nota_final is not None and nota_final >= 0:
            notas_finales.append(nota_final)

    if not notas_finales:
        return {
            "total_alumnos": len(alumnos_grupo),
            "total_con_nota": 0,
            "aprobados": 0,
            "desaprobados": 0,
            "tasa_aprobacion": 0,
            "nota_promedio": 0,
            "nota_maxima": 0,
            "nota_minima": 0,
            "alumno_maxima": None,
            "alumno_minima": None,
        }

    # Calcular estadísticas
    aprobados = sum(1 for nota in notas_finales if nota >= 10.5)
    desaprobados = len(notas_finales) - aprobados

    return {
        "total_alumnos": len(alumnos_grupo),
        "total_con_nota": len(notas_finales),
        "aprobados": aprobados,
        "desaprobados": desaprobados,
        "tasa_aprobacion": round((aprobados / len(notas_finales) * 100), 2),
        "nota_promedio": round(sum(notas_finales) / len(notas_finales), 2),
        "nota_maxima": max(notas_finales),
        "nota_minima": min(notas_finales),
        "alumno_maxima": obtener_alumno_nota_maxima(grupo_teoria, alumnos_grupo),
        "alumno_minima": obtener_alumno_nota_minima(grupo_teoria, alumnos_grupo),
    }


def calcular_estadisticas_detalladas(grupo_teoria, alumnos_grupo):
    """
    Calcula estadísticas detalladas por tipo de evaluación - VERSIÓN CON -1
    """
    curso = grupo_teoria.curso
    estadisticas = {"parciales": {}, "continuas": {}, "sustitutorios": {}}

    # Estadísticas de parciales
    for periodo in [1, 2, 3]:
        peso_parcial = getattr(curso, f"peso_parcial_{periodo}", 0)
        if peso_parcial > 0:
            notas_parcial = []
            alumnos_con_nota = 0
            alumnos_evaluados = 0

            for matricula in alumnos_grupo:
                nota = Nota.objects.filter(
                    alumno=matricula.alumno, curso=curso, tipo="P", periodo=periodo
                ).first()
                # Solo considerar notas válidas (>= 0)
                if nota and nota.valor is not None and nota.valor >= 0:
                    notas_parcial.append(nota.valor)
                    alumnos_con_nota += 1
                # Contar alumnos que tienen registro (aunque sea -1)
                if nota:
                    alumnos_evaluados += 1

            if notas_parcial:
                estadisticas["parciales"][f"P{periodo}"] = {
                    "promedio": round(sum(notas_parcial) / len(notas_parcial), 2),
                    "maxima": max(notas_parcial),
                    "minima": min(notas_parcial),
                    "total": len(notas_parcial),
                    "alumnos_con_nota": alumnos_con_nota,
                    "alumnos_evaluados": alumnos_evaluados,
                    "alumnos_sin_nota": len(alumnos_grupo) - alumnos_con_nota,
                    "peso": peso_parcial,
                    "tiene_datos": len(notas_parcial) > 0,
                }
            else:
                # Incluir período incluso sin notas válidas
                estadisticas["parciales"][f"P{periodo}"] = {
                    "promedio": 0,
                    "maxima": 0,
                    "minima": 0,
                    "total": 0,
                    "alumnos_con_nota": 0,
                    "alumnos_evaluados": alumnos_evaluados,
                    "alumnos_sin_nota": len(alumnos_grupo),
                    "peso": peso_parcial,
                    "tiene_datos": False,
                }

    # Estadísticas de continuas
    for periodo in [1, 2, 3]:
        peso_continua = getattr(curso, f"peso_continua_{periodo}", 0)
        if peso_continua > 0:
            notas_continua = []
            alumnos_con_nota = 0
            alumnos_evaluados = 0

            for matricula in alumnos_grupo:
                nota = Nota.objects.filter(
                    alumno=matricula.alumno, curso=curso, tipo="C", periodo=periodo
                ).first()
                # Solo considerar notas válidas (>= 0)
                if nota and nota.valor is not None and nota.valor >= 0:
                    notas_continua.append(nota.valor)
                    alumnos_con_nota += 1
                # Contar alumnos que tienen registro (aunque sea -1)
                if nota:
                    alumnos_evaluados += 1

            if notas_continua:
                estadisticas["continuas"][f"C{periodo}"] = {
                    "promedio": round(sum(notas_continua) / len(notas_continua), 2),
                    "maxima": max(notas_continua),
                    "minima": min(notas_continua),
                    "total": len(notas_continua),
                    "alumnos_con_nota": alumnos_con_nota,
                    "alumnos_evaluados": alumnos_evaluados,
                    "alumnos_sin_nota": len(alumnos_grupo) - alumnos_con_nota,
                    "peso": peso_continua,
                    "tiene_datos": len(notas_continua) > 0,
                }
            else:
                # Incluir período incluso sin notas válidas
                estadisticas["continuas"][f"C{periodo}"] = {
                    "promedio": 0,
                    "maxima": 0,
                    "minima": 0,
                    "total": 0,
                    "alumnos_con_nota": 0,
                    "alumnos_evaluados": alumnos_evaluados,
                    "alumnos_sin_nota": len(alumnos_grupo),
                    "peso": peso_continua,
                    "tiene_datos": False,
                }

    # Estadísticas de sustitutorios
    notas_sustitutorio = []
    alumnos_con_sustitutorio = 0
    alumnos_evaluados_sust = 0

    for matricula in alumnos_grupo:
        nota = Nota.objects.filter(
            alumno=matricula.alumno, curso=curso, tipo="S"
        ).first()
        # Solo considerar notas válidas (>= 0)
        if nota and nota.valor is not None and nota.valor >= 0:
            notas_sustitutorio.append(nota.valor)
            alumnos_con_sustitutorio += 1
        # Contar alumnos que tienen registro (aunque sea -1)
        if nota:
            alumnos_evaluados_sust += 1

    if notas_sustitutorio:
        estadisticas["sustitutorios"] = {
            "promedio": round(sum(notas_sustitutorio) / len(notas_sustitutorio), 2),
            "maxima": max(notas_sustitutorio),
            "minima": min(notas_sustitutorio),
            "total": len(notas_sustitutorio),
            "alumnos_con_nota": alumnos_con_sustitutorio,
            "alumnos_evaluados": alumnos_evaluados_sust,
            "alumnos_sin_nota": len(alumnos_grupo) - alumnos_con_sustitutorio,
            "tiene_datos": True,
        }
    else:
        estadisticas["sustitutorios"] = {
            "promedio": 0,
            "maxima": 0,
            "minima": 0,
            "total": 0,
            "alumnos_con_nota": 0,
            "alumnos_evaluados": alumnos_evaluados_sust,
            "alumnos_sin_nota": len(alumnos_grupo),
            "tiene_datos": False,
        }

    return estadisticas


def obtener_alumno_nota_maxima(grupo_teoria, alumnos_grupo):
    """Obtiene el alumno con la nota más alta - VERSIÓN CON -1"""
    curso = grupo_teoria.curso
    mejor_alumno = None
    mejor_nota = -1  # Iniciar con -1

    for matricula in alumnos_grupo:
        nota_final = calcular_nota_final(matricula.alumno, curso)
        # Solo considerar notas válidas (>= 0)
        if nota_final is not None and nota_final >= 0:
            if mejor_nota == -1 or nota_final > mejor_nota:
                mejor_nota = nota_final
                mejor_alumno = matricula.alumno

    return {"alumno": mejor_alumno, "nota": mejor_nota} if mejor_alumno else None


def obtener_alumno_nota_minima(grupo_teoria, alumnos_grupo):
    """Obtiene el alumno con la nota más baja - VERSIÓN CON -1"""
    curso = grupo_teoria.curso
    peor_alumno = None
    peor_nota = -1  # Iniciar con -1

    for matricula in alumnos_grupo:
        nota_final = calcular_nota_final(matricula.alumno, curso)
        # Solo considerar notas válidas (>= 0)
        if nota_final is not None and nota_final >= 0:
            if peor_nota == -1 or nota_final < peor_nota:
                peor_nota = nota_final
                peor_alumno = matricula.alumno

    return {"alumno": peor_alumno, "nota": peor_nota} if peor_alumno else None


def procesar_examenes(request, grupo_teoria):
    """
    Procesa la subida de exámenes PDF usando datos de la base de datos
    """
    archivos = request.FILES.getlist("archivos_examen")
    tipo_examen = request.POST.get("tipo_examen")

    # Mapeo de tipos de examen para el nombre del archivo
    tipo_map = {"A": "alta", "P": "promedio", "B": "baja"}

    tipo_nombre = tipo_map.get(tipo_examen, "examen")
    curso_codigo = grupo_teoria.curso.codigo or "0000"

    archivos_procesados = 0
    errores = []

    # Obtener todos los alumnos del grupo desde la base de datos
    alumnos_grupo = MatriculaCurso.objects.filter(
        curso=grupo_teoria.curso, turno=grupo_teoria.turno
    ).select_related("alumno")

    # Crear diccionario de alumnos por DNI para búsqueda rápida
    alumnos_dict = {
        alumno_matricula.alumno.dni: alumno_matricula.alumno
        for alumno_matricula in alumnos_grupo
    }

    for archivo in archivos:
        # Extraer DNI del nombre del archivo
        nombre_archivo = archivo.name

        # Buscar DNI en el nombre del archivo (puede estar en diferentes formatos)
        import re

        dni_match = re.search(r"(\d{8})", nombre_archivo)

        if not dni_match:
            errores.append(f"Formato inválido: {nombre_archivo} - No se encontró DNI")
            continue

        dni_alumno = dni_match.group(1)

        # Buscar alumno en el diccionario
        alumno = alumnos_dict.get(dni_alumno)

        if not alumno:
            errores.append(f"Alumno con DNI {dni_alumno} no encontrado en este grupo")
            continue

        try:
            # Generar nuevo nombre para el archivo usando datos de la base de datos
            nuevo_nombre = f"{alumno.dni}_{curso_codigo}_{tipo_nombre}.pdf"

            # Asignar el nuevo nombre al archivo
            archivo.name = nuevo_nombre

            # Crear o actualizar examen
            examen, created = Examen.objects.get_or_create(
                alumno=alumno,
                GrupoTeoria=grupo_teoria,
                tipo=tipo_examen,
                defaults={"archivo": archivo},
            )

            if not created:
                # Si ya existe, eliminar el archivo anterior y guardar el nuevo
                if examen.archivo:
                    examen.archivo.delete(save=False)
                examen.archivo = archivo
                examen.save()

            archivos_procesados += 1

        except Exception as e:
            errores.append(f"Error con {nombre_archivo}: {str(e)}")
            continue

    # Mostrar mensajes de resultado
    if archivos_procesados > 0:
        messages.success(
            request, f"Se procesaron {archivos_procesados} archivos correctamente"
        )

    if errores:
        messages.error(request, f"Errores: {', '.join(errores[:5])}")


def descargar_estadisticas_excel(request, grupo_teoria, alumnos_grupo):
    """
    Genera y descarga un Excel con las estadísticas - VERSIÓN CON -1
    """
    try:
        curso = grupo_teoria.curso

        # Crear DataFrame con datos de alumnos
        data = []
        for matricula in alumnos_grupo:
            alumno = matricula.alumno
            nota_final = calcular_nota_final(alumno, curso)

            # Obtener notas individuales (solo mostrar >= 0)
            notas_parciales = {1: "", 2: "", 3: ""}
            notas_continuas = {1: "", 2: "", 3: ""}
            nota_sustitutorio = ""

            for periodo in [1, 2, 3]:
                nota_parcial = Nota.objects.filter(
                    alumno=alumno, curso=curso, tipo="P", periodo=periodo
                ).first()
                if (
                    nota_parcial
                    and nota_parcial.valor is not None
                    and nota_parcial.valor >= 0
                ):
                    notas_parciales[periodo] = nota_parcial.valor

                nota_continua = Nota.objects.filter(
                    alumno=alumno, curso=curso, tipo="C", periodo=periodo
                ).first()
                if (
                    nota_continua
                    and nota_continua.valor is not None
                    and nota_continua.valor >= 0
                ):
                    notas_continuas[periodo] = nota_continua.valor

            nota_sust = Nota.objects.filter(
                alumno=alumno, curso=curso, tipo="S"
            ).first()
            if nota_sust and nota_sust.valor is not None and nota_sust.valor >= 0:
                nota_sustitutorio = nota_sust.valor

            # Determinar estado
            estado = "Sin nota"
            if nota_final is not None and nota_final >= 0:
                estado = "Aprobado" if nota_final >= 10.5 else "Desaprobado"

            data.append(
                {
                    "DNI": alumno.dni,
                    "Alumno": alumno.nombre,
                    "Parcial 1": notas_parciales[1],
                    "Parcial 2": notas_parciales[2],
                    "Parcial 3": notas_parciales[3],
                    "Continua 1": notas_continuas[1],
                    "Continua 2": notas_continuas[2],
                    "Continua 3": notas_continuas[3],
                    "Sustitutorio": nota_sustitutorio,
                    "Nota Final": nota_final
                    if nota_final is not None and nota_final >= 0
                    else "",
                    "Estado": estado,
                }
            )

        # Crear DataFrame principal
        df = pd.DataFrame(data)

        # Crear output
        output = io.BytesIO()

        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            # Hoja 1: Datos de alumnos
            df.to_excel(writer, sheet_name="Estadisticas_Alumnos", index=False)

            # Hoja 2: Resumen general
            resumen_data = []
            if data:
                # Solo considerar notas válidas (>= 0)
                notas_finales = [
                    d["Nota Final"]
                    for d in data
                    if d["Nota Final"] != "" and float(d["Nota Final"]) >= 0
                ]

                if notas_finales:
                    aprobados = sum(1 for d in data if d["Estado"] == "Aprobado")
                    desaprobados = sum(1 for d in data if d["Estado"] == "Desaprobado")
                    sin_nota = sum(1 for d in data if d["Estado"] == "Sin nota")

                    resumen_data.extend(
                        [
                            ["Total Alumnos", len(data)],
                            ["Con Nota Final", len(notas_finales)],
                            ["Sin Nota Final", sin_nota],
                            ["Aprobados", aprobados],
                            ["Desaprobados", desaprobados],
                            [
                                "Tasa Aprobación",
                                f"{(aprobados / len(notas_finales) * 100):.2f}%"
                                if notas_finales
                                else "0%",
                            ],
                            [
                                "Nota Promedio",
                                f"{sum(notas_finales) / len(notas_finales):.2f}"
                                if notas_finales
                                else "0.00",
                            ],
                            [
                                "Nota Máxima",
                                f"{max(notas_finales):.2f}"
                                if notas_finales
                                else "0.00",
                            ],
                            [
                                "Nota Mínima",
                                f"{min(notas_finales):.2f}"
                                if notas_finales
                                else "0.00",
                            ],
                        ]
                    )
                else:
                    resumen_data.extend(
                        [
                            ["Total Alumnos", len(data)],
                            ["Con Nota Final", 0],
                            ["Sin Nota Final", len(data)],
                            ["Aprobados", 0],
                            ["Desaprobados", 0],
                            ["Tasa Aprobación", "0%"],
                            ["Nota Promedio", "0.00"],
                            ["Nota Máxima", "0.00"],
                            ["Nota Mínima", "0.00"],
                        ]
                    )

            df_resumen = pd.DataFrame(resumen_data, columns=["Métrica", "Valor"])
            df_resumen.to_excel(writer, sheet_name="Resumen_General", index=False)

            # Ajustar el ancho de las columnas automáticamente
            for sheet_name in writer.sheets:
                worksheet = writer.sheets[sheet_name]
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[column_letter].width = adjusted_width

        output.seek(0)

        # Crear respuesta HTTP
        response = HttpResponse(
            output.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = (
            f'attachment; filename="estadisticas_{grupo_teoria.curso.nombre}_{grupo_teoria.turno}.xlsx"'
        )

        return response

    except Exception as e:
        messages.error(request, f"Error al generar Excel: {str(e)}")
        return redirect("revisar_estadisticas")


def calcular_estadisticas_avance(grupo_teoria, alumnos_grupo):
    """
    Calcula estadísticas de avance por periodos para gráficos - VERSIÓN CON -1
    """
    curso = grupo_teoria.curso
    estadisticas_avance = {
        "parciales": {
            "periodos": [],
            "promedios": [],
            "tiene_datos": [],
            "alumnos_evaluados": [],
        },
        "continuas": {
            "periodos": [],
            "promedios": [],
            "tiene_datos": [],
            "alumnos_evaluados": [],
        },
    }

    # Estadísticas de avance de parciales
    for periodo in [1, 2, 3]:
        peso_parcial = getattr(curso, f"peso_parcial_{periodo}", 0)
        if peso_parcial > 0:
            notas_periodo = []
            alumnos_evaluados = 0

            for matricula in alumnos_grupo:
                nota = Nota.objects.filter(
                    alumno=matricula.alumno, curso=curso, tipo="P", periodo=periodo
                ).first()
                # Solo considerar notas válidas (>= 0)
                if nota and nota.valor is not None and nota.valor >= 0:
                    notas_periodo.append(nota.valor)
                    alumnos_evaluados += 1

            tiene_datos = len(notas_periodo) > 0
            promedio = (
                round(sum(notas_periodo) / len(notas_periodo), 2) if tiene_datos else 0
            )

            estadisticas_avance["parciales"]["periodos"].append(f"P{periodo}")
            estadisticas_avance["parciales"]["promedios"].append(promedio)
            estadisticas_avance["parciales"]["tiene_datos"].append(tiene_datos)
            estadisticas_avance["parciales"]["alumnos_evaluados"].append(
                alumnos_evaluados
            )

    # Estadísticas de avance de continuas
    for periodo in [1, 2, 3]:
        peso_continua = getattr(curso, f"peso_continua_{periodo}", 0)
        if peso_continua > 0:
            notas_periodo = []
            alumnos_evaluados = 0

            for matricula in alumnos_grupo:
                nota = Nota.objects.filter(
                    alumno=matricula.alumno, curso=curso, tipo="C", periodo=periodo
                ).first()
                # Solo considerar notas válidas (>= 0)
                if nota and nota.valor is not None and nota.valor >= 0:
                    notas_periodo.append(nota.valor)
                    alumnos_evaluados += 1

            tiene_datos = len(notas_periodo) > 0
            promedio = (
                round(sum(notas_periodo) / len(notas_periodo), 2) if tiene_datos else 0
            )

            estadisticas_avance["continuas"]["periodos"].append(f"C{periodo}")
            estadisticas_avance["continuas"]["promedios"].append(promedio)
            estadisticas_avance["continuas"]["tiene_datos"].append(tiene_datos)
            estadisticas_avance["continuas"]["alumnos_evaluados"].append(
                alumnos_evaluados
            )

    return estadisticas_avance


# =======================Vista de Admin====================================================


def inicio_admin(request):
    return render(request, "siscad/admin/menu.html")
