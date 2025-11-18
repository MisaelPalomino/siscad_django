from ..comunes.imports import *
def ingresar_notas(request):
    if "email" not in request.session:
        return redirect("login")

    profesor = get_object_or_404(Profesor, email=request.session["email"])
    grupos_teoria = GrupoTeoria.objects.filter(profesor=profesor)

    grupo_seleccionado = None
    alumnos_grupo = []
    notas_data = []

    if request.method == "POST":
        grupo_id = request.POST.get("grupo_id")

        if grupo_id:
            grupo_seleccionado = get_object_or_404(
                GrupoTeoria, id=grupo_id, profesor=profesor
            )

            alumnos_grupo = MatriculaCurso.objects.filter(
                curso=grupo_seleccionado.curso, turno=grupo_seleccionado.turno
            ).select_related("alumno")

            if "guardar_manual" in request.POST:
                procesar_notas_manual(request, grupo_seleccionado, alumnos_grupo)
                messages.success(request, "Notas guardadas correctamente")
                return redirect("ingresar_notas")

            elif "procesar_excel" in request.POST and request.FILES.get(
                "archivo_excel"
            ):
                archivo = request.FILES["archivo_excel"]
                resultado = procesar_excel_notas(archivo, grupo_seleccionado)
                if resultado["success"]:
                    messages.success(request, resultado["message"])
                    return redirect("ingresar_notas")
                else:
                    messages.error(request, resultado["message"])

            notas_data = preparar_datos_notas(alumnos_grupo, grupo_seleccionado)

    context = {
        "grupos_teoria": grupos_teoria,
        "grupo_seleccionado": grupo_seleccionado,
        "alumnos_grupo": alumnos_grupo,
        "notas_data": notas_data,
    }

    return render(request, "siscad/profesor/ingresar_notas.html", context)


def procesar_notas_manual(request, grupo_teoria, alumnos_grupo):
    """
    Procesa las notas ingresadas manualmente en el formulario
    """
    curso = grupo_teoria.curso

    with transaction.atomic():
        for alumno_matricula in alumnos_grupo:
            alumno = alumno_matricula.alumno

            # Procesar notas continuas (solo si el curso las tiene configuradas)
            if any(
                [
                    curso.peso_continua_1 > 0,
                    curso.peso_continua_2 > 0,
                    curso.peso_continua_3 > 0,
                ]
            ):
                for periodo in [1, 2, 3]:
                    peso_continua = getattr(curso, f"peso_continua_{periodo}", 0)
                    # Solo procesar si el periodo tiene peso configurado
                    if peso_continua > 0:
                        campo_continua = f"continua_{alumno.id}_{periodo}"
                        valor_continua = request.POST.get(campo_continua)

                        if valor_continua and valor_continua.strip():
                            try:
                                valor_float = float(valor_continua)
                                nota, created = Nota.objects.get_or_create(
                                    alumno=alumno,
                                    curso=curso,
                                    tipo="C",
                                    periodo=periodo,
                                    defaults={
                                        "valor": valor_float,
                                        "peso": peso_continua,
                                    },
                                )
                                if not created:
                                    nota.valor = valor_float
                                    nota.peso = peso_continua
                                    nota.save()
                            except ValueError:
                                continue
                        # Si el campo está vacío, establecer como -1
                        elif valor_continua == "":
                            Nota.objects.filter(
                                alumno=alumno, curso=curso, tipo="C", periodo=periodo
                            ).update(valor=-1)

            # Procesar 3 parciales (solo si el curso los tiene configurados)
            if any(
                [
                    curso.peso_parcial_1 > 0,
                    curso.peso_parcial_2 > 0,
                    curso.peso_parcial_3 > 0,
                ]
            ):
                for periodo_parcial in [1, 2, 3]:
                    peso_parcial = getattr(curso, f"peso_parcial_{periodo_parcial}", 0)
                    # Solo procesar si el periodo tiene peso configurado
                    if peso_parcial > 0:
                        campo_parcial = f"parcial_{alumno.id}_{periodo_parcial}"
                        valor_parcial = request.POST.get(campo_parcial)

                        if valor_parcial and valor_parcial.strip():
                            try:
                                valor_float = float(valor_parcial)
                                nota, created = Nota.objects.get_or_create(
                                    alumno=alumno,
                                    curso=curso,
                                    tipo="P",
                                    periodo=periodo_parcial,
                                    defaults={
                                        "valor": valor_float,
                                        "peso": peso_parcial,
                                    },
                                )
                                if not created:
                                    nota.valor = valor_float
                                    nota.peso = peso_parcial
                                    nota.save()
                            except ValueError:
                                continue
                        # Si el campo está vacío, establecer como -1
                        elif valor_parcial == "":
                            Nota.objects.filter(
                                alumno=alumno,
                                curso=curso,
                                tipo="P",
                                periodo=periodo_parcial,
                            ).update(valor=-1)

            # Procesar sustitutorio (solo si el curso tiene al menos 2 parciales)
            if curso.peso_parcial_1 > 0 and curso.peso_parcial_2 > 0:
                campo_sustitutorio = f"sustitutorio_{alumno.id}"
                valor_sustitutorio = request.POST.get(campo_sustitutorio)

                if valor_sustitutorio and valor_sustitutorio.strip():
                    try:
                        valor_float = float(valor_sustitutorio)
                        nota, created = Nota.objects.get_or_create(
                            alumno=alumno,
                            curso=curso,
                            tipo="S",
                            periodo=1,
                            defaults={"valor": valor_float, "peso": 1},
                        )
                        if not created:
                            nota.valor = valor_float
                            nota.save()
                    except ValueError:
                        continue
                # Si el campo está vacío, establecer como -1
                elif valor_sustitutorio == "":
                    Nota.objects.filter(
                        alumno=alumno, curso=curso, tipo="S", periodo=1
                    ).update(valor=-1)


def procesar_excel_notas(archivo, grupo_teoria):
    """
    Procesa un archivo Excel para cargar notas masivamente - VERSIÓN CON -1
    """
    try:
        # Leer el archivo más rápido sin procesamiento extra
        if archivo.name.endswith(".xlsx"):
            df = pd.read_excel(archivo, dtype={"dni_alumno": str, "valor": object})
        elif archivo.name.endswith(".csv"):
            df = pd.read_csv(archivo, dtype={"dni_alumno": str, "valor": object})
        else:
            return {"success": False, "message": "Formato de archivo no soportado"}

        # Validar columnas rápidamente
        columnas_requeridas = ["dni_alumno", "tipo_nota", "periodo", "valor"]
        if not all(col in df.columns for col in columnas_requeridas):
            return {"success": False, "message": "Faltan columnas requeridas"}

        # Limpiar y preparar datos
        df = df.copy()
        df["dni_alumno"] = df["dni_alumno"].astype(str).str.strip()
        df["tipo_nota"] = df["tipo_nota"].astype(str).str.strip().str.upper()
        df["periodo"] = (
            pd.to_numeric(df["periodo"], errors="coerce").fillna(0).astype(int)
        )

        # Manejar valores vacíos: convertir a -1
        df["valor"] = pd.to_numeric(df["valor"], errors="coerce")
        df["valor"] = df["valor"].fillna(-1)  # Valores vacíos se convierten a -1

        # Filtrar valores válidos (>= 0 y <= 20) y -1 para "no asignado"
        df = df[((df["valor"] >= 0) & (df["valor"] <= 20)) | (df["valor"] == -1)]

        # Filtrar tipos de nota válidos y periodos válidos
        df = df[df["tipo_nota"].isin(["C", "P", "S"]) & df["periodo"].isin([1, 2, 3])]

        curso = grupo_teoria.curso

        # Pre-cache de datos para mejor performance
        alumnos_dni = set(df["dni_alumno"].unique())
        alumnos_dict = {
            alumno.dni: alumno for alumno in Alumno.objects.filter(dni__in=alumnos_dni)
        }

        # Verificar matriculas en una sola consulta
        matriculas_existentes = set(
            MatriculaCurso.objects.filter(
                alumno__dni__in=alumnos_dni, curso=curso, turno=grupo_teoria.turno
            ).values_list("alumno__dni", flat=True)
        )

        # Filtrar solo alumnos matriculados
        df = df[df["dni_alumno"].isin(matriculas_existentes)]

        if df.empty:
            return {
                "success": False,
                "message": "No hay datos válidos para procesar",
            }

        # Preparar datos para bulk_create y bulk_update
        notas_a_procesar = []

        # Agrupar por alumno y tipo para procesamiento más eficiente
        for (dni_alumno, tipo_nota, periodo), group_df in df.groupby(
            ["dni_alumno", "tipo_nota", "periodo"]
        ):
            if not group_df.empty:
                valor = group_df["valor"].iloc[0]  # Tomar el primer valor

                # Validar configuración del curso
                if tipo_nota == "C":
                    peso = getattr(curso, f"peso_continua_{periodo}", 0)
                    if peso == 0:
                        continue
                elif tipo_nota == "P":
                    peso = getattr(curso, f"peso_parcial_{periodo}", 0)
                    if peso == 0:
                        continue
                else:  # Sustitutorio
                    if not (curso.peso_parcial_1 > 0 and curso.peso_parcial_2 > 0):
                        continue
                    peso = 1

                alumno = alumnos_dict.get(dni_alumno)
                if not alumno:
                    continue

                notas_a_procesar.append(
                    {
                        "alumno": alumno,
                        "tipo": tipo_nota,
                        "periodo": periodo,
                        "valor": valor,
                        "peso": peso,
                    }
                )

        if not notas_a_procesar:
            return {"success": False, "message": "No hay notas válidas para procesar"}

        # Procesamiento masivo con bulk operations
        with transaction.atomic():
            # Obtener notas existentes para evitar duplicados
            notas_existentes = Nota.objects.filter(
                alumno__in=[n["alumno"] for n in notas_a_procesar], curso=curso
            ).select_related("alumno")

            # Crear diccionario de notas existentes para búsqueda rápida
            existentes_dict = {}
            for nota in notas_existentes:
                key = (nota.alumno.dni, nota.tipo, nota.periodo)
                existentes_dict[key] = nota

            # Separar en crear y actualizar
            notas_para_crear = []
            notas_para_actualizar = []

            for nota_data in notas_a_procesar:
                key = (nota_data["alumno"].dni, nota_data["tipo"], nota_data["periodo"])
                if key in existentes_dict:
                    nota_existente = existentes_dict[key]
                    nota_existente.valor = nota_data["valor"]
                    nota_existente.peso = nota_data["peso"]
                    notas_para_actualizar.append(nota_existente)
                else:
                    notas_para_crear.append(
                        Nota(
                            alumno=nota_data["alumno"],
                            curso=curso,
                            tipo=nota_data["tipo"],
                            periodo=nota_data["periodo"],
                            valor=nota_data["valor"],
                            peso=nota_data["peso"],
                        )
                    )

            # Ejecutar operaciones masivas
            if notas_para_crear:
                Nota.objects.bulk_create(notas_para_crear, batch_size=1000)

            if notas_para_actualizar:
                Nota.objects.bulk_update(
                    notas_para_actualizar, ["valor", "peso"], batch_size=1000
                )

            total_procesadas = len(notas_para_crear) + len(notas_para_actualizar)

        return {
            "success": True,
            "message": f"Se procesaron {total_procesadas} registros correctamente "
            f"({len(notas_para_crear)} nuevos, {len(notas_para_actualizar)} actualizados)",
        }

    except Exception as e:
        return {"success": False, "message": f"Error al procesar el archivo: {str(e)}"}


def preparar_datos_notas(alumnos_grupo, grupo_teoria):
    """
    Prepara los datos de notas para mostrar en la tabla
    """
    datos = []
    curso = grupo_teoria.curso

    for matricula in alumnos_grupo:
        alumno = matricula.alumno

        # Obtener todas las notas del alumno en este curso
        notas_alumno = Nota.objects.filter(alumno=alumno, curso=curso)

        # Organizar notas (-1 se muestra como vacío)
        notas_continua = {1: None, 2: None, 3: None}
        notas_parcial = {1: None, 2: None, 3: None}
        nota_sustitutorio = None

        for nota in notas_alumno:
            if nota.tipo == "C" and nota.periodo in [1, 2, 3]:
                # Mostrar solo valores >= 0, -1 se muestra como None/vacío
                notas_continua[nota.periodo] = nota.valor if nota.valor >= 0 else None
            elif nota.tipo == "P" and nota.periodo in [1, 2, 3]:
                notas_parcial[nota.periodo] = nota.valor if nota.valor >= 0 else None
            elif nota.tipo == "S":
                nota_sustitutorio = nota.valor if nota.valor >= 0 else None

        # Determinar qué tipos de notas mostrar según configuración del curso
        curso_tiene_continua = any(
            [
                curso.peso_continua_1 > 0,
                curso.peso_continua_2 > 0,
                curso.peso_continua_3 > 0,
            ]
        )

        curso_tiene_parciales = any(
            [
                curso.peso_parcial_1 > 0,
                curso.peso_parcial_2 > 0,
                curso.peso_parcial_3 > 0,
            ]
        )

        curso_tiene_sustitutorio = curso.peso_parcial_1 > 0 and curso.peso_parcial_2 > 0

        datos.append(
            {
                "alumno": alumno,
                "matricula": matricula,
                "notas_continua": notas_continua,
                "notas_parcial": notas_parcial,
                "nota_sustitutorio": nota_sustitutorio,
                "curso_tiene_continua": curso_tiene_continua,
                "curso_tiene_parciales": curso_tiene_parciales,
                "curso_tiene_sustitutorio": curso_tiene_sustitutorio,
                "pesos_continua": {
                    1: curso.peso_continua_1,
                    2: curso.peso_continua_2,
                    3: curso.peso_continua_3,
                },
                "pesos_parcial": {
                    1: curso.peso_parcial_1,
                    2: curso.peso_parcial_2,
                    3: curso.peso_parcial_3,
                },
            }
        )

    return datos


def descargar_plantilla_excel(request, grupo_id):
    """
    Vista para descargar una plantilla Excel para cargar notas CON DATOS EXISTENTES
    """
    grupo_teoria = get_object_or_404(GrupoTeoria, id=grupo_id)
    curso = grupo_teoria.curso

    # Crear DataFrame con la estructura requerida
    data = {
        "dni_alumno": [],
        "tipo_nota": [],  # C: Continua, P: Parcial, S: Sustitutorio
        "periodo": [],  # 1, 2, 3 para continua/parcial, 1 para sustitutorio
        "valor": [],
    }

    # Agregar alumnos del grupo
    alumnos_grupo = MatriculaCurso.objects.filter(
        curso=grupo_teoria.curso, turno=grupo_teoria.turno
    ).select_related("alumno")

    # Pre-cargar todas las notas de estos alumnos en este curso para mejor performance
    alumnos_ids = [matricula.alumno.id for matricula in alumnos_grupo]
    notas_existentes = Nota.objects.filter(
        alumno_id__in=alumnos_ids, curso=curso
    ).select_related("alumno")

    # Organizar notas por alumno para acceso rápido
    notas_por_alumno = {}
    for nota in notas_existentes:
        if nota.alumno_id not in notas_por_alumno:
            notas_por_alumno[nota.alumno_id] = []
        notas_por_alumno[nota.alumno_id].append(nota)

    for matricula in alumnos_grupo:
        alumno = matricula.alumno
        notas_alumno = notas_por_alumno.get(alumno.id, [])

        # Organizar notas del alumno por tipo y periodo
        notas_organizadas = {
            "C": {1: None, 2: None, 3: None},
            "P": {1: None, 2: None, 3: None},
            "S": {1: None},
        }

        for nota in notas_alumno:
            if nota.tipo in ["C", "P"] and nota.periodo in [1, 2, 3]:
                # Solo mostrar valores >= 0, -1 se muestra como vacío
                notas_organizadas[nota.tipo][nota.periodo] = (
                    nota.valor if nota.valor >= 0 else None
                )
            elif nota.tipo == "S" and nota.periodo == 1:
                notas_organizadas["S"][1] = nota.valor if nota.valor >= 0 else None

        # Agregar filas para notas continuas (solo si el curso las tiene)
        if any(
            [
                curso.peso_continua_1 > 0,
                curso.peso_continua_2 > 0,
                curso.peso_continua_3 > 0,
            ]
        ):
            for periodo in [1, 2, 3]:
                peso_continua = getattr(curso, f"peso_continua_{periodo}", 0)
                if peso_continua > 0:
                    data["dni_alumno"].append(alumno.dni)
                    data["tipo_nota"].append("C")
                    data["periodo"].append(periodo)
                    # Usar valor existente si existe y es >= 0, sino vacío
                    valor_existente = notas_organizadas["C"][periodo]
                    data["valor"].append(
                        valor_existente if valor_existente is not None else ""
                    )

        # Agregar filas para notas parciales (solo si el curso las tiene)
        if any(
            [
                curso.peso_parcial_1 > 0,
                curso.peso_parcial_2 > 0,
                curso.peso_parcial_3 > 0,
            ]
        ):
            for periodo in [1, 2, 3]:
                peso_parcial = getattr(curso, f"peso_parcial_{periodo}", 0)
                if peso_parcial > 0:
                    data["dni_alumno"].append(alumno.dni)
                    data["tipo_nota"].append("P")
                    data["periodo"].append(periodo)
                    # Usar valor existente si existe y es >= 0, sino vacío
                    valor_existente = notas_organizadas["P"][periodo]
                    data["valor"].append(
                        valor_existente if valor_existente is not None else ""
                    )

        # Agregar fila para sustitutorio (solo si el curso lo permite)
        if curso.peso_parcial_1 > 0 and curso.peso_parcial_2 > 0:
            data["dni_alumno"].append(alumno.dni)
            data["tipo_nota"].append("S")
            data["periodo"].append(1)
            # Usar valor existente si existe y es >= 0, sino vacío
            valor_existente = notas_organizadas["S"][1]
            data["valor"].append(valor_existente if valor_existente is not None else "")

    df = pd.DataFrame(data)

    # Crear un libro de Excel con formato mejorado
    output = io.BytesIO()

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        # Hoja principal con datos
        df.to_excel(writer, sheet_name="Plantilla_Notas", index=False)

        # Obtener la hoja para aplicar formato
        workbook = writer.book
        worksheet = writer.sheets["Plantilla_Notas"]

        # Aplicar formato a las columnas
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        # Estilo para encabezados
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_fill = PatternFill(
            start_color="366092", end_color="366092", fill_type="solid"
        )
        header_alignment = Alignment(horizontal="center", vertical="center")
        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # Aplicar estilo a los encabezados
        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border

        # Aplicar bordes a todas las celdas con datos
        for row in worksheet.iter_rows(
            min_row=2, max_row=worksheet.max_row, min_col=1, max_col=4
        ):
            for cell in row:
                cell.border = border

        # Ajustar ancho de columnas
        column_widths = {
            "A": 15,  # dni_alumno
            "B": 12,  # tipo_nota
            "C": 10,  # periodo
            "D": 12,  # valor
        }

        for col, width in column_widths.items():
            worksheet.column_dimensions[col].width = width

        # Aplicar formato de texto a las columnas
        for row in range(2, worksheet.max_row + 1):
            # DNI como texto
            worksheet[f"A{row}"].number_format = "@"
            # Valor con formato numérico (2 decimales)
            worksheet[f"D{row}"].number_format = "0.00"

        # Hoja de instrucciones
        instrucciones_data = {
            "Instrucciones": [
                "1. NO modifique la estructura de las columnas",
                "2. Los tipos de nota son: C=Continua, P=Parcial, S=Sustitutorio",
                "3. Los periodos válidos son: 1, 2, 3",
                "4. Las notas deben estar entre 0 y 20",
                "5. Deje vacío si no desea modificar una nota existente",
                "6. Para eliminar una nota, déjela vacía en el Excel",
                "7. Mantenga el formato del DNI sin espacios",
                "8. Los valores vacíos se interpretarán como 'no evaluado'",
                "9. Solo se procesarán las notas que coincidan con la configuración del curso",
                "10. El sustitutorio solo aplica si el curso tiene Parcial 1 y Parcial 2 configurados",
            ]
        }
        df_instrucciones = pd.DataFrame(instrucciones_data)
        df_instrucciones.to_excel(writer, sheet_name="Instrucciones", index=False)

        # Formato hoja de instrucciones
        worksheet_inst = writer.sheets["Instrucciones"]

        # Aplicar formato a la hoja de instrucciones
        for cell in worksheet_inst[1]:
            cell.font = Font(bold=True, color="FFFFFF", size=12)
            cell.fill = PatternFill(
                start_color="28a745", end_color="28a745", fill_type="solid"
            )
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border

        # Ajustar ancho de columna de instrucciones
        worksheet_inst.column_dimensions["A"].width = 80

        # Aplicar bordes y formato a las instrucciones
        for row in range(2, worksheet_inst.max_row + 1):
            worksheet_inst[f"A{row}"].border = border
            worksheet_inst[f"A{row}"].alignment = Alignment(
                wrap_text=True, vertical="center"
            )

        # Hoja de configuración del curso
        config_data = {
            "Configuración del Curso": [
                f"Curso: {curso.nombre}",
                f"Código: {curso.codigo}",
                f"Grupo: {grupo_teoria.get_turno_display()}",
                "",
                "Pesos de Evaluación:",
                f"Continua P1: {curso.peso_continua_1}%",
                f"Continua P2: {curso.peso_continua_2}%",
                f"Continua P3: {curso.peso_continua_3}%",
                f"Parcial 1: {curso.peso_parcial_1}%",
                f"Parcial 2: {curso.peso_parcial_2}%",
                f"Parcial 3: {curso.peso_parcial_3}%",
                "",
                "Sustitutorio: "
                + (
                    "HABILITADO (reemplaza la nota más baja entre P1 y P2)"
                    if curso.peso_parcial_1 > 0 and curso.peso_parcial_2 > 0
                    else "NO HABILITADO"
                ),
                "",
                f"Total alumnos en el grupo: {len(alumnos_grupo)}",
                f"Fecha de generación: {timezone.now().strftime('%Y-%m-%d %H:%M')}",
            ]
        }
        df_config = pd.DataFrame(config_data)
        df_config.to_excel(writer, sheet_name="Configuración", index=False)

        # Formato hoja de configuración
        worksheet_config = writer.sheets["Configuración"]
        worksheet_config.column_dimensions["A"].width = 60

        # Aplicar formato a la hoja de configuración
        for cell in worksheet_config[1]:
            cell.font = Font(bold=True, color="FFFFFF", size=12)
            cell.fill = PatternFill(
                start_color="6f42c1", end_color="6f42c1", fill_type="solid"
            )
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border

        # Aplicar bordes y formato a la configuración
        for row in range(2, worksheet_config.max_row + 1):
            worksheet_config[f"A{row}"].border = border
            worksheet_config[f"A{row}"].alignment = Alignment(
                wrap_text=True, vertical="center"
            )

        # Resaltar filas importantes en configuración
        important_rows = [5, 12]  # Filas de "Pesos de Evaluación" y "Sustitutorio"
        for row in important_rows:
            if row <= worksheet_config.max_row:
                worksheet_config[f"A{row}"].font = Font(bold=True, color="2c3e50")

    output.seek(0)

    # Crear respuesta HTTP
    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = (
        f'attachment; filename="plantilla_notas_{curso.codigo}_{grupo_teoria.get_turno_display()}_{timezone.now().strftime("%Y%m%d")}.xlsx"'
    )

    return response


