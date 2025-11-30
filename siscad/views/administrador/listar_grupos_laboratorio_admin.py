from ..comunes.imports import *
import openpyxl
from openpyxl.utils import get_column_letter


def listar_grupos_laboratorio_admin(request):
    laboratorios = GrupoLaboratorio.objects.select_related(
        "grupo_teoria", "grupo_teoria__curso", "profesor"
    ).all()

    if request.method == "POST":
        # Actualizar cupos de un laboratorio
        if "actualizar_cupos" in request.POST:
            lab_id = request.POST.get("lab_id")
            nuevos_cupos = request.POST.get("cupos")

            if lab_id and nuevos_cupos:
                try:
                    lab = GrupoLaboratorio.objects.get(id=lab_id)
                    lab.cupos = int(nuevos_cupos)
                    lab.save()
                    messages.success(
                        request,
                        f"Cupos actualizados para {lab.grupo_teoria.curso.nombre} - Lab {lab.grupo}",
                    )
                    return redirect("listar_grupos_laboratorio_admin")
                except GrupoLaboratorio.DoesNotExist:
                    messages.error(request, "Laboratorio no encontrado")
                except ValueError:
                    messages.error(request, "El número de cupos debe ser un entero")

        # Actualizar profesor de un laboratorio
        elif "actualizar_profesor" in request.POST:
            lab_id = request.POST.get("lab_id")
            profesor_id = request.POST.get("profesor_id")

            if lab_id and profesor_id:
                try:
                    lab = GrupoLaboratorio.objects.get(id=lab_id)
                    profesor = Profesor.objects.get(id=profesor_id)
                    lab.profesor = profesor
                    lab.save()
                    messages.success(
                        request,
                        f"Profesor actualizado para {lab.grupo_teoria.curso.nombre} - Lab {lab.grupo}",
                    )
                    return redirect("listar_grupos_laboratorio_admin")
                except GrupoLaboratorio.DoesNotExist:
                    messages.error(request, "Laboratorio no encontrado")
                except Profesor.DoesNotExist:
                    messages.error(request, "Profesor no encontrado")

        # Descargar Excel
        elif "descargar_excel" in request.POST:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Laboratorios"

            headers = [
                "Curso",
                "Código Curso",
                "Grupo Teoría",
                "Grupo Laboratorio",
                "Profesor",
                "Email Profesor",
                "Cupos",
                "Matriculados",
            ]
            for col_num, header in enumerate(headers, 1):
                ws[f"{get_column_letter(col_num)}1"] = header

            for row_num, lab in enumerate(laboratorios, 2):
                # Contar matriculados en este laboratorio
                matriculados = MatriculaLaboratorio.objects.filter(
                    grupo_laboratorio=lab
                ).count()

                ws[f"A{row_num}"] = lab.grupo_teoria.curso.nombre
                ws[f"B{row_num}"] = lab.grupo_teoria.curso.codigo or "N/A"
                ws[f"C{row_num}"] = lab.grupo_teoria.turno
                ws[f"D{row_num}"] = lab.grupo
                ws[f"E{row_num}"] = (
                    lab.profesor.nombre if lab.profesor else "Sin asignar"
                )
                ws[f"F{row_num}"] = lab.profesor.email if lab.profesor else "N/A"
                ws[f"G{row_num}"] = lab.cupos
                ws[f"H{row_num}"] = matriculados

            # Ajustar el ancho de las columnas
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

            archivo_nombre = f"Laboratorios_Disponibles_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
            response = HttpResponse(
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"    
            )
            response["Content-Disposition"] = f'attachment; filename="{archivo_nombre}"'
            wb.save(response)
            return response

    # Obtener todos los profesores para los dropdowns
    profesores = Profesor.objects.all().order_by("nombre")

    # Calcular matriculados para cada laboratorio


    for lab in laboratorios:
        lab.matriculados = MatriculaLaboratorio.objects.filter(
            grupo_laboratorio=lab
        ).count()
        lab.disponibles = max(
            0, lab.cupos
        )  # Asegurar que no sea negativo
        lab.porcentaje_ocupacion = (
            (lab.matriculados / (lab.cupos + lab.matriculados) * 100) if lab.cupos > 0 else 0
        )
        lab.total = lab.cupos + lab.matriculados

    return render(
        request,
        "siscad/admin/listar_grupos_laboratorio_admin.html",
        {
            "laboratorios": laboratorios,
            "profesores": profesores,
        },
    )
