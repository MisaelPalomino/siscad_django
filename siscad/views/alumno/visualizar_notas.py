from ..comunes.imports import *


def visualizar_notas(request):
    rol = request.session.get("rol")

    if rol != "Alumno":
        request.session["rol"] = "Ninguno"
        return redirect("login")

    nombre = request.session.get("nombre")

    alumno = Alumno.objects.filter(nombre=nombre).first()

    if not alumno:
        return redirect("inicio_alumno")

    # Verificar si se solicita descargar Excel
    if request.method == "POST" and "descargar_excel" in request.POST:
        return descargar_libreta_excel(alumno)

    matriculas = MatriculaCurso.objects.filter(alumno=alumno).select_related("curso")

    notas_por_curso = []
    estadisticas_generales = {
        "total_cursos": 0,
        "cursos_aprobados": 0,
        "cursos_desaprobados": 0,
        "cursos_en_proceso": 0,
        "promedio_general": 0,
        "mejor_nota": 0,
        "peor_nota": 20,
        "total_notas_registradas": 0,
        "cursos_con_sustitutorio": 0,
        "total_notas_esperadas": 0,
        "progreso_general": 0,
    }

    suma_promedios = 0
    cursos_con_promedio = 0
    total_notas_registradas_global = 0
    total_notas_esperadas_global = 0

    for matricula in matriculas:
        curso = matricula.curso
        turno = matricula.turno

        # Obtener todas las notas del curso
        notas_curso = Nota.objects.filter(alumno=alumno, curso=curso).order_by(
            "tipo", "periodo"
        )

        # Organizar notas por tipo - VERSIÓN CON -1
        notas_parciales = {1: None, 2: None, 3: None}
        notas_continuas = {1: None, 2: None, 3: None}
        nota_sustitutorio = None

        for nota in notas_curso:
            if nota.tipo == "P" and nota.periodo in [1, 2, 3]:
                # Solo considerar valores >= 0
                notas_parciales[nota.periodo] = (
                    nota.valor if nota.valor is not None and nota.valor >= 0 else None
                )
            elif nota.tipo == "C" and nota.periodo in [1, 2, 3]:
                notas_continuas[nota.periodo] = (
                    nota.valor if nota.valor is not None and nota.valor >= 0 else None
                )
            elif nota.tipo == "S":
                nota_sustitutorio = (
                    nota.valor if nota.valor is not None and nota.valor >= 0 else None
                )

        # Calcular nota final considerando sustitutorio
        promedio_final = calcular_nota_final_alumno(alumno, curso)

        # Determinar estado del curso
        estado_curso = determinar_estado_curso(
            promedio_final, curso, notas_parciales, notas_continuas
        )

        # Estadísticas del curso - VERSIÓN CON -1
        notas_validas = [
            n.valor for n in notas_curso if n.valor is not None and n.valor >= 0
        ]
        total_notas_curso = len(notas_validas)
        promedio_curso = sum(notas_validas) / len(notas_validas) if notas_validas else 0

        # Calcular progreso del curso
        total_notas_esperadas = calcular_total_notas_esperadas(curso)
        progreso_curso = (
            (total_notas_curso / total_notas_esperadas * 100)
            if total_notas_esperadas > 0
            else 0
        )

        # Actualizar estadísticas generales
        estadisticas_generales["total_cursos"] += 1
        estadisticas_generales["total_notas_registradas"] += total_notas_curso
        total_notas_registradas_global += total_notas_curso
        total_notas_esperadas_global += total_notas_esperadas

        if estado_curso == "aprobado":
            estadisticas_generales["cursos_aprobados"] += 1
        elif estado_curso == "desaprobado":
            estadisticas_generales["cursos_desaprobados"] += 1
        else:
            estadisticas_generales["cursos_en_proceso"] += 1

        if promedio_final and promedio_final >= 0:
            suma_promedios += promedio_final
            cursos_con_promedio += 1
            estadisticas_generales["mejor_nota"] = max(
                estadisticas_generales["mejor_nota"], promedio_final
            )
            estadisticas_generales["peor_nota"] = min(
                estadisticas_generales["peor_nota"], promedio_final
            )

        if nota_sustitutorio is not None and nota_sustitutorio >= 0:
            estadisticas_generales["cursos_con_sustitutorio"] += 1

        # Preparar datos para mostrar en tabla
        datos_parciales = []
        datos_continuas = []

        for periodo in [1, 2, 3]:
            peso_parcial = getattr(curso, f"peso_parcial_{periodo}", 0)
            if peso_parcial > 0:
                datos_parciales.append(
                    {
                        "periodo": periodo,
                        "nota": notas_parciales[periodo],
                        "peso": peso_parcial,
                        "tiene_nota": notas_parciales[periodo] is not None
                        and notas_parciales[periodo] >= 0,
                    }
                )

            peso_continua = getattr(curso, f"peso_continua_{periodo}", 0)
            if peso_continua > 0:
                datos_continuas.append(
                    {
                        "periodo": periodo,
                        "nota": notas_continuas[periodo],
                        "peso": peso_continua,
                        "tiene_nota": notas_continuas[periodo] is not None
                        and notas_continuas[periodo] >= 0,
                    }
                )

        notas_por_curso.append(
            {
                "curso": curso,
                "nombre_curso": curso.nombre,
                "codigo_curso": curso.codigo,
                "turno": turno,
                "datos_parciales": datos_parciales,
                "datos_continuas": datos_continuas,
                "nota_sustitutorio": nota_sustitutorio,
                "tiene_sustitutorio": nota_sustitutorio is not None
                and nota_sustitutorio >= 0,
                "promedio_final": round(promedio_final, 2)
                if promedio_final is not None and promedio_final >= 0
                else None,
                "estado": estado_curso,
                "progreso": round(progreso_curso, 1),
                "total_notas_registradas": total_notas_curso,
                "total_notas_esperadas": total_notas_esperadas,
                "notas_totales": list(notas_curso),
                "configuracion_curso": {
                    "tiene_continua": any(
                        [
                            curso.peso_continua_1 > 0,
                            curso.peso_continua_2 > 0,
                            curso.peso_continua_3 > 0,
                        ]
                    ),
                    "tiene_parciales": any(
                        [
                            curso.peso_parcial_1 > 0,
                            curso.peso_parcial_2 > 0,
                            curso.peso_parcial_3 > 0,
                        ]
                    ),
                    "tiene_sustitutorio": (
                        curso.peso_parcial_1 > 0 and curso.peso_parcial_2 > 0
                    ),
                    "pesos_continua": {
                        1: curso.peso_continua_1,
                        2: curso.peso_continua_2,
                        3: curso.peso_continua_3,
                    },
                    "pesos_parcial": {
                        1: curso.peso_parcial_1,
                        2: curso.peso_parcial_2,
                        3: curso.peso_parcial_3,
                    },
                },
            }
        )

    # Calcular promedio general
    if cursos_con_promedio > 0:
        estadisticas_generales["promedio_general"] = round(
            suma_promedios / cursos_con_promedio, 2
        )

    # Calcular progreso general
    if total_notas_esperadas_global > 0:
        estadisticas_generales["progreso_general"] = round(
            (total_notas_registradas_global / total_notas_esperadas_global) * 100, 1
        )
    else:
        estadisticas_generales["progreso_general"] = 0

    # Calcular porcentajes para gráficos
    if estadisticas_generales["total_cursos"] > 0:
        estadisticas_generales["porcentaje_aprobados"] = round(
            (
                estadisticas_generales["cursos_aprobados"]
                / estadisticas_generales["total_cursos"]
            )
            * 100,
            1,
        )
        estadisticas_generales["porcentaje_desaprobados"] = round(
            (
                estadisticas_generales["cursos_desaprobados"]
                / estadisticas_generales["total_cursos"]
            )
            * 100,
            1,
        )
        estadisticas_generales["porcentaje_en_proceso"] = round(
            (
                estadisticas_generales["cursos_en_proceso"]
                / estadisticas_generales["total_cursos"]
            )
            * 100,
            1,
        )
    else:
        estadisticas_generales.update(
            {
                "porcentaje_aprobados": 0,
                "porcentaje_desaprobados": 0,
                "porcentaje_en_proceso": 0,
            }
        )

    # Ordenar cursos por estado y promedio
    notas_por_curso.sort(
        key=lambda x: (
            0 if x["estado"] == "aprobado" else 1 if x["estado"] == "en_proceso" else 2,
            -x["promedio_final"] if x["promedio_final"] else 0,
        )
    )

    context = {
        "alumno": alumno,
        "notas_por_curso": notas_por_curso,
        "estadisticas": estadisticas_generales,
        "semestre_actual": alumno.calcular_semestre() or "No asignado",
        "nombre": nombre,
        "rol": rol,
    }

    return render(request, "siscad/alumno/visualizar_notas.html", context)


def descargar_libreta_excel(alumno):
    """
    Genera y descarga un Excel con la libreta de notas del alumno
    """
    try:
        import pandas as pd
        import io
        from django.http import HttpResponse
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from django.utils import timezone

        # Obtener todas las matrículas del alumno
        matriculas = MatriculaCurso.objects.filter(alumno=alumno).select_related(
            "curso"
        )

        # Crear datos para el Excel
        data = []

        for matricula in matriculas:
            curso = matricula.curso
            notas_curso = Nota.objects.filter(alumno=alumno, curso=curso).order_by(
                "tipo", "periodo"
            )

            # Organizar notas
            notas_parciales = {1: "", 2: "", 3: ""}
            notas_continuas = {1: "", 2: "", 3: ""}
            nota_sustitutorio = ""

            for nota in notas_curso:
                if nota.tipo == "P" and nota.periodo in [1, 2, 3]:
                    if nota.valor is not None and nota.valor >= 0:
                        notas_parciales[nota.periodo] = nota.valor
                elif nota.tipo == "C" and nota.periodo in [1, 2, 3]:
                    if nota.valor is not None and nota.valor >= 0:
                        notas_continuas[nota.periodo] = nota.valor
                elif nota.tipo == "S":
                    if nota.valor is not None and nota.valor >= 0:
                        nota_sustitutorio = nota.valor

            # Calcular nota final
            nota_final = calcular_nota_final_alumno(alumno, curso)

            # Determinar estado
            estado = "En proceso"
            if nota_final is not None and nota_final >= 0:
                estado = "Aprobado" if nota_final >= 10.5 else "Desaprobado"

            data.append(
                {
                    "Código Curso": curso.codigo,
                    "Nombre Curso": curso.nombre,
                    "Turno": matricula.get_turno_display(),
                    "Continua 1": notas_continuas[1],
                    "Continua 2": notas_continuas[2],
                    "Continua 3": notas_continuas[3],
                    "Parcial 1": notas_parciales[1],
                    "Parcial 2": notas_parciales[2],
                    "Parcial 3": notas_parciales[3],
                    "Sustitutorio": nota_sustitutorio,
                    "Nota Final": nota_final
                    if nota_final is not None and nota_final >= 0
                    else "",
                    "Estado": estado,
                }
            )

        # Crear DataFrame
        df = pd.DataFrame(data)

        # Crear output
        output = io.BytesIO()

        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            # Hoja principal con datos
            df.to_excel(writer, sheet_name="Libreta de Notas", index=False)

            # Obtener la hoja para aplicar formato
            workbook = writer.book
            worksheet = writer.sheets["Libreta de Notas"]

            # Aplicar formato profesional
            header_font = Font(bold=True, color="FFFFFF", size=12)
            header_fill = PatternFill(
                start_color="366092", end_color="366092", fill_type="solid"
            )
            header_alignment = Alignment(horizontal="center", vertical="center")
            border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )

            # Aplicar estilo a los encabezados
            for cell in worksheet[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = border

            # Aplicar bordes a todas las celdas
            for row in worksheet.iter_rows(
                min_row=2, max_row=worksheet.max_row, min_col=1, max_col=len(df.columns)
            ):
                for cell in row:
                    cell.border = border

            # Ajustar ancho de columnas
            column_widths = {
                "A": 12,
                "B": 30,
                "C": 10,
                "D": 10,
                "E": 10,
                "F": 10,
                "G": 10,
                "H": 10,
                "I": 10,
                "J": 12,
                "K": 12,
                "L": 12,
            }

            for col, width in column_widths.items():
                worksheet.column_dimensions[col].width = width

            # Aplicar formato condicional para estados
            for row in range(2, worksheet.max_row + 1):
                estado_cell = worksheet[f"L{row}"]
                if estado_cell.value == "Aprobado":
                    estado_cell.fill = PatternFill(
                        start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"
                    )
                    estado_cell.font = Font(color="006100", bold=True)
                elif estado_cell.value == "Desaprobado":
                    estado_cell.fill = PatternFill(
                        start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"
                    )
                    estado_cell.font = Font(color="9C0006", bold=True)
                else:
                    estado_cell.fill = PatternFill(
                        start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"
                    )
                    estado_cell.font = Font(color="9C6500", bold=True)

            # Hoja de resumen
            resumen_data = [
                ["ALUMNO", f"{alumno.nombre}"],
                ["DNI", f"{alumno.dni}"],
                ["SEMESTRE", f"{alumno.calcular_semestre() or 'No asignado'}"],
                ["FECHA DE GENERACIÓN", f"{timezone.now().strftime('%Y-%m-%d %H:%M')}"],
                ["", ""],
                ["RESUMEN ACADÉMICO", ""],
                ["Total de Cursos", len(data)],
                [
                    "Cursos Aprobados",
                    len([d for d in data if d["Estado"] == "Aprobado"]),
                ],
                [
                    "Cursos Desaprobados",
                    len([d for d in data if d["Estado"] == "Desaprobado"]),
                ],
                [
                    "Cursos en Proceso",
                    len([d for d in data if d["Estado"] == "En proceso"]),
                ],
                ["Promedio General", f"{calcular_promedio_general(alumno):.2f}"],
            ]

            df_resumen = pd.DataFrame(resumen_data, columns=["Concepto", "Valor"])
            df_resumen.to_excel(writer, sheet_name="Resumen", index=False)

            # Formato hoja de resumen
            worksheet_resumen = writer.sheets["Resumen"]
            worksheet_resumen.column_dimensions["A"].width = 25
            worksheet_resumen.column_dimensions["B"].width = 20

            # Aplicar formato a la hoja de resumen
            for cell in worksheet_resumen[1]:
                cell.font = Font(bold=True, color="FFFFFF", size=12)
                cell.fill = PatternFill(
                    start_color="28a745", end_color="28a745", fill_type="solid"
                )
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = border

            for row in range(2, worksheet_resumen.max_row + 1):
                for col in ["A", "B"]:
                    cell = worksheet_resumen[f"{col}{row}"]
                    cell.border = border
                    if row >= 7:  # Filas del resumen académico
                        cell.font = Font(bold=True)

        output.seek(0)

        # Crear respuesta HTTP
        response = HttpResponse(
            output.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = (
            f'attachment; filename="libreta_notas_{alumno.dni}_{timezone.now().strftime("%Y%m%d")}.xlsx"'
        )

        return response

    except Exception as e:
        # En un entorno real, deberías manejar este error adecuadamente
        from django.contrib import messages

        return redirect("visualizar_notas")


def calcular_promedio_general(alumno):
    """
    Calcula el promedio general del alumno
    """
    matriculas = MatriculaCurso.objects.filter(alumno=alumno).select_related("curso")
    suma_promedios = 0
    cursos_con_promedio = 0

    for matricula in matriculas:
        curso = matricula.curso
        nota_final = calcular_nota_final_alumno(alumno, curso)
        if nota_final is not None and nota_final >= 0:
            suma_promedios += nota_final
            cursos_con_promedio += 1

    return suma_promedios / cursos_con_promedio if cursos_con_promedio > 0 else 0


# Funciones auxiliares actualizadas para trabajar con -1
def calcular_nota_final_alumno(alumno, curso):
    """
    Calcula la nota final para un alumno en un curso específico - VERSIÓN CON -1
    """
    notas = Nota.objects.filter(alumno=alumno, curso=curso)

    # Solo considerar notas válidas (>= 0)
    continuas = {
        n.periodo: n
        for n in notas
        if n.tipo == "C" and n.valor is not None and n.valor >= 0
    }
    parciales = {
        n.periodo: n
        for n in notas
        if n.tipo == "P" and n.valor is not None and n.valor >= 0
    }
    sustitutorio = next(
        (n for n in notas if n.tipo == "S" and n.valor is not None and n.valor >= 0),
        None,
    )

    # Aplicar sustitutorio si existe
    if (
        sustitutorio
        and sustitutorio.valor is not None
        and sustitutorio.valor >= 0
        and 1 in parciales
        and 2 in parciales
    ):
        if parciales[1].valor <= parciales[2].valor:
            nota_original_p1 = parciales[1].valor
            parciales[1].valor = sustitutorio.valor
        else:
            nota_original_p2 = parciales[2].valor
            parciales[2].valor = sustitutorio.valor

    # Calcular promedio ponderado
    total_puntos = 0
    total_pesos = 0

    # Sumar continuas
    for periodo in [1, 2, 3]:
        if periodo in continuas:
            peso = getattr(curso, f"peso_continua_{periodo}", 0)
            if peso > 0:
                total_puntos += continuas[periodo].valor * peso
                total_pesos += peso

    # Sumar parciales
    for periodo in [1, 2, 3]:
        if periodo in parciales:
            peso = getattr(curso, f"peso_parcial_{periodo}", 0)
            if peso > 0:
                total_puntos += parciales[periodo].valor * peso
                total_pesos += peso

    if total_pesos == 0:
        return None

    nota_final = total_puntos / total_pesos

    # Restaurar notas originales si se aplicó sustitutorio
    if (
        sustitutorio
        and sustitutorio.valor is not None
        and sustitutorio.valor >= 0
        and 1 in parciales
        and 2 in parciales
    ):
        if "nota_original_p1" in locals():
            parciales[1].valor = nota_original_p1
        elif "nota_original_p2" in locals():
            parciales[2].valor = nota_original_p2

    return round(nota_final, 2)


def determinar_estado_curso(promedio_final, curso, notas_parciales, notas_continuas):
    """
    Determina el estado del curso (aprobado, desaprobado, en proceso) - VERSIÓN CON -1
    """
    if promedio_final is None or promedio_final < 0:
        return "en_proceso"

    # Verificar si todas las notas necesarias están registradas
    notas_faltantes = False

    # Verificar parciales
    for periodo in [1, 2, 3]:
        peso_parcial = getattr(curso, f"peso_parcial_{periodo}", 0)
        if peso_parcial > 0 and (
            notas_parciales[periodo] is None or notas_parciales[periodo] < 0
        ):
            notas_faltantes = True

    # Verificar continuas
    for periodo in [1, 2, 3]:
        peso_continua = getattr(curso, f"peso_continua_{periodo}", 0)
        if peso_continua > 0 and (
            notas_continuas[periodo] is None or notas_continuas[periodo] < 0
        ):
            notas_faltantes = True

    if notas_faltantes:
        return "en_proceso"

    # Determinar aprobación basado en el promedio final
    if promedio_final >= 10.5:  # Umbral de aprobación
        return "aprobado"
    else:
        return "desaprobado"


def calcular_total_notas_esperadas(curso):
    """
    Calcula el total de notas esperadas para un curso según su configuración
    """
    total = 0

    # Contar parciales esperados
    for periodo in [1, 2, 3]:
        if getattr(curso, f"peso_parcial_{periodo}", 0) > 0:
            total += 1

    # Contar continuas esperadas
    for periodo in [1, 2, 3]:
        if getattr(curso, f"peso_continua_{periodo}", 0) > 0:
            total += 1

    # Agregar sustitutorio si aplica
    if curso.peso_parcial_1 > 0 and curso.peso_parcial_2 > 0:
        total += 1

    return total
